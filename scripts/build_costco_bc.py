"""Build BC sheets - Costco.xlsx — Yuno Payment Orchestrator Business Case."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/germantatis/Desktop/GTMCoding/Business Cases/BC sheets - Costco.xlsx"

# ------ INPUTS ------
ANNUAL_VISITS = 1_267_000_000  # 1.267B
CONVERSION = 0.06
INDUSTRY = "Retail (Pets, Electronics...)"
TAKE_RATE = 0.15  # Costco gross margin proxy (low margin wholesale)
LOCAL_MDR = 0.018
LOCAL_MDR_REDUCTION = -0.002
LOCAL_MDR_NEW = LOCAL_MDR + LOCAL_MDR_REDUCTION
DELTA_ACCEPTANCE_LOCAL = 0.03
APMS_MDR_COST = 0.012
COST_PER_MONTH = 0.0216  # $21,600/month
TIME_PER_APM = 3
TIME_PER_PSP = 4
YUNO_INTEGRATION = 2
VTC = 0.22
CTS = 0.85

# Per-country inputs
countries = {
    "United States": {
        "region": "NA", "sow": 0.6015, "atv": 160, "local_entity": "Yes",
        "current_apms_list": "Costco Anywhere Visa (Citi), Affirm BNPL, Costco Shop Card",
        "current_apms_count": 1,
        "sow_cards": 0.92, "sow_apms": 0.08,
        "apms_to_integrate": "Apple Pay, Google Pay, PayPal, Venmo",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.04,
        "pct_tpv_shift": 0.02,
    },
    "Canada": {
        "region": "NA", "sow": 0.2022, "atv": 145, "local_entity": "Yes",
        "current_apms_list": "CIBC Costco Mastercard, Interac (in-warehouse), CIBC Pace It",
        "current_apms_count": 2,
        "sow_cards": 0.88, "sow_apms": 0.12,
        "apms_to_integrate": "PayPal, Apple Pay, Google Pay, Interac e-Transfer",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.06,
        "pct_tpv_shift": 0.03,
    },
    "Mexico": {
        "region": "LATAM", "sow": 0.0514, "atv": 100, "local_entity": "Yes",
        "current_apms_list": "Costco Banamex Card, PayPal, MSI installments (credit cards)",
        "current_apms_count": 2,
        "sow_cards": 0.85, "sow_apms": 0.15,
        "apms_to_integrate": "OXXO Pay, SPEI, Mercado Pago, Spin by OXXO",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.10,
        "pct_tpv_shift": 0.05,
    },
    "United Kingdom": {
        "region": "EMEA", "sow": 0.0351, "atv": 135, "local_entity": "Yes",
        "current_apms_list": "Visa, Mastercard (no PayPal online per UK guides)",
        "current_apms_count": 0,
        "sow_cards": 0.95, "sow_apms": 0.05,
        "apms_to_integrate": "PayPal, Apple Pay, Google Pay, Open Banking",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.05,
        "pct_tpv_shift": 0.025,
    },
    "Taiwan": {
        "region": "APAC", "sow": 0.0326, "atv": 130, "local_entity": "Yes",
        "current_apms_list": "Fubon Costco Card, cash (non-Costco intl cards not accepted)",
        "current_apms_count": 1,
        "sow_cards": 0.90, "sow_apms": 0.10,
        "apms_to_integrate": "LINE Pay, JKO Pay, Apple Pay, ibon Konbini",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.08,
        "pct_tpv_shift": 0.04,
    },
    "Japan": {
        "region": "APAC", "sow": 0.0228, "atv": 115, "local_entity": "Yes",
        "current_apms_list": "Mastercard only, Orico Costco Global Card, cash",
        "current_apms_count": 1,
        "sow_cards": 0.93, "sow_apms": 0.07,
        "apms_to_integrate": "PayPay, Rakuten Pay, LINE Pay, Konbini (7-Eleven/FamilyMart)",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.10,
        "pct_tpv_shift": 0.05,
    },
    "Australia": {
        "region": "APAC", "sow": 0.0192, "atv": 135, "local_entity": "Yes",
        "current_apms_list": "NOT VERIFIED (page blocked); assume cards-first per global pattern",
        "current_apms_count": 1,
        "sow_cards": 0.90, "sow_apms": 0.10,
        "apms_to_integrate": "PayPal, Apple Pay, Google Pay, Afterpay",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.06,
        "pct_tpv_shift": 0.03,
    },
    "South Korea": {
        "region": "APAC", "sow": 0.0176, "atv": 130, "local_entity": "Yes",
        "current_apms_list": "Hyundai Card (exclusive), Apple Pay (since Mar 2023)",
        "current_apms_count": 2,
        "sow_cards": 0.88, "sow_apms": 0.12,
        "apms_to_integrate": "KakaoPay, NaverPay, Toss, Samsung Pay",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.08,
        "pct_tpv_shift": 0.04,
    },
}

# Calculate per-country metrics
for c, d in countries.items():
    d["visits"] = ANNUAL_VISITS * d["sow"]
    d["txs"] = d["visits"] * CONVERSION
    d["tpv"] = d["txs"] * d["atv"] / 1_000_000  # in $M
    d["tpv_cards"] = d["tpv"] * d["sow_cards"]
    d["tpv_apms"] = d["tpv"] * d["sow_apms"]
    d["delta_ar"] = DELTA_ACCEPTANCE_LOCAL
    d["incr_tpv"] = d["tpv_cards"] * d["delta_ar"]
    d["incr_rev"] = d["incr_tpv"] * TAKE_RATE
    d["annual_cards_tpv"] = d["tpv_cards"] + d["incr_tpv"]
    d["current_mdr"] = LOCAL_MDR
    d["mdr_reduction"] = LOCAL_MDR_REDUCTION
    d["new_mdr"] = LOCAL_MDR_NEW
    d["mdr_savings"] = d["annual_cards_tpv"] * abs(LOCAL_MDR_REDUCTION)
    d["cards_benefit"] = d["incr_rev"] + d["mdr_savings"]
    d["apms_incr_tpv"] = d["tpv"] * d["pct_incremental_tx"]
    d["apms_incr_rev"] = d["apms_incr_tpv"] * TAKE_RATE
    d["apms_shifted_tpv"] = d["tpv"] * d["pct_tpv_shift"]
    d["apm_mdr_savings"] = d["apms_shifted_tpv"] * (d["current_mdr"] - APMS_MDR_COST)
    d["apms_benefit"] = d["apms_incr_rev"] + d["apm_mdr_savings"]
    d["total_benefit"] = d["cards_benefit"] + d["apms_benefit"]
    # Eng
    d["apm_int_months"] = d["apms_count_int"] * TIME_PER_APM
    d["apm_int_cost"] = d["apm_int_months"] * COST_PER_MONTH
    d["mor_count"] = 1  # all have local entity
    d["psp_int_months"] = d["mor_count"] * TIME_PER_PSP
    d["psp_int_cost"] = d["psp_int_months"] * COST_PER_MONTH
    d["total_int_months"] = d["apm_int_months"] + d["psp_int_months"]
    d["eng_savings"] = d["apm_int_cost"] + d["psp_int_cost"]

# Regions ordered
regions = {
    "NA": ["United States", "Canada"],
    "LATAM": ["Mexico"],
    "EMEA": ["United Kingdom"],
    "APAC": ["Taiwan", "Japan", "Australia", "South Korea"],
}

# Totals
tot_acceptance_uplift = sum(d["incr_rev"] for d in countries.values())
tot_fee_renegotiation = sum(d["mdr_savings"] for d in countries.values())
tot_new_apm_growth = sum(d["apms_benefit"] for d in countries.values())
tot_eng_savings = sum(d["eng_savings"] for d in countries.values())
tot_benefit = tot_acceptance_uplift + tot_fee_renegotiation + tot_new_apm_growth + tot_eng_savings

tot_int_months = sum(d["total_int_months"] for d in countries.values())
yuno_total_int_months = YUNO_INTEGRATION * len(regions)  # one Yuno integration per region
time_saved = tot_int_months - yuno_total_int_months
time_saved_years = time_saved / 12

# ------ EXCEL BUILD ------
wb = openpyxl.Workbook()
del wb["Sheet"]

# Styles
purple_fill = PatternFill(start_color="6B2CB0", end_color="6B2CB0", fill_type="solid")
header_fill = PatternFill(start_color="EFE6FA", end_color="EFE6FA", fill_type="solid")
total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=12)
bold_font = Font(bold=True, size=11)
section_font = Font(bold=True, size=11, color="6B2CB0")
title_font = Font(bold=True, size=16, color="FFFFFF")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def title_row(ws, txt, ncols):
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=2, value=txt)
    cell.font = title_font
    cell.fill = purple_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

def write_row(ws, row, vals, bold=False, fill=None, fmt=None, indent_col=2):
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=indent_col + i, value=v)
        if bold:
            cell.font = bold_font
        if fill:
            cell.fill = fill
        if fmt and isinstance(v, (int, float)) and i > 0:
            cell.number_format = fmt
        cell.border = border

def section_header(ws, row, txt, ncols):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=2, value=txt)
    c.font = section_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="center")

# ===== RESUMEN =====
ws = wb.create_sheet("RESUMEN")
title_row(ws, "BUSINESS CASE — COSTCO WHOLESALE", 8)

r = 3
write_row(ws, r, ["Company", "Costco Wholesale Corporation"], bold=True); r += 1
write_row(ws, r, ["Industry", INDUSTRY], bold=True); r += 1
write_row(ws, r, ["Annual Total Visits", ANNUAL_VISITS], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Annual Revenue ($M)", 275235], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Take Rate", TAKE_RATE], bold=True, fmt="0.00%"); r += 2

section_header(ws, r, "TOTAL BENEFIT SUMMARY", 8); r += 1
write_row(ws, r, ["Category", "Value ($M)", "Lower Range ($M)", "Upper Range ($M)"], bold=True, fill=total_fill); r += 1
def lo_hi(v):
    return [round(v, 2), round(v * 0.85, 2), round(v * 1.15, 2)]
write_row(ws, r, ["TPV Acceptance Rate Uplift"] + lo_hi(tot_acceptance_uplift), fmt="#,##0.00"); r += 1
write_row(ws, r, ["Fee Renegotiation (MDR Savings)"] + lo_hi(tot_fee_renegotiation), fmt="#,##0.00"); r += 1
write_row(ws, r, ["New APM TPV Growth"] + lo_hi(tot_new_apm_growth), fmt="#,##0.00"); r += 1
write_row(ws, r, ["Engineering Cost Savings"] + lo_hi(tot_eng_savings), fmt="#,##0.00"); r += 1
write_row(ws, r, ["TOTAL BENEFIT"] + lo_hi(tot_benefit), bold=True, fill=total_fill, fmt="#,##0.00"); r += 2

write_row(ws, r, ["Total integration months (without Yuno)", tot_int_months], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Yuno integration time (months, per region x4)", yuno_total_int_months], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Time Saved (months)", time_saved], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Time Saved (years)", round(time_saved_years, 1)], bold=True, fmt="0.0"); r += 2

problems = [
    "Single-network exclusive deals per country (Visa-only US, Mastercard-only Canada/UK, Hyundai-only Korea, Fubon-only Taiwan, Banamex-only Mexico, Orico-only Japan) create acquirer lock-in with no failover",
    "No wallets on Costco.com globally: Apple Pay quietly removed from Costco.com and app in May 2025; no Google Pay, no PayPal in US/UK; PayPal exists only in Mexico online",
    "BNPL is US-only via Affirm exclusive (May 2025 launch); Canada, UK, Mexico, Australia, APAC have zero Costco BNPL coverage despite local BNPL penetration (Afterpay AU, Klarna UK, Tabby Tamara MENA equivalents)",
    "APM gaps in every market: no OXXO/SPEI/Mercado Pago in Mexico, no Interac e-Transfer in Canada, no Open Banking in UK, no PayPay/Konbini in Japan, no LINE Pay/JKO in Taiwan, no KakaoPay/NaverPay in Korea",
    "Citi Visa exclusive co-brand contract expires in 2026 (renewal options through 2029); single-PSP renewal exposure with no orchestration optionality to leverage alternative acquirers",
]
for i, p in enumerate(problems, 1):
    write_row(ws, r, [None, None, None, None, None, f"PROBLEM {i}", p], indent_col=2); r += 1

# Column widths
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 4
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 110

# ===== PAY-INS PER REGION =====
def build_payins_sheet(region_name, country_list):
    sheet_name = f"BC Pay-ins {region_name}"
    ws = wb.create_sheet(sheet_name)
    ncols = 3 + len(country_list)  # label + countries + total
    title_row(ws, f"BC PAY-INS — {region_name.upper()}", ncols)

    headers = ["SDR INPUTS"] + country_list + ["TOTAL"]

    rows_data = []
    rows_data.append(("section", "TRAFFIC & TPV"))
    rows_data.append(("data", "Web Traffic (SoW)", [countries[c]["sow"] for c in country_list], "0.00%", None))
    rows_data.append(("data", "Web Traffic Visits", [int(countries[c]["visits"]) for c in country_list], "#,##0", None))
    rows_data.append(("data", "Conversion", [CONVERSION] * len(country_list), "0.00%", None))
    rows_data.append(("data", "View to Checkout (VtC)", [VTC] * len(country_list), "0.00%", None))
    rows_data.append(("data", "Checkout to Success (CtS)", [CTS] * len(country_list), "0.00%", None))
    rows_data.append(("data", "Successful Transactions/Year", [int(countries[c]["txs"]) for c in country_list], "#,##0", "sum"))
    rows_data.append(("data", "ATV ($)", [countries[c]["atv"] for c in country_list], "$#,##0", None))
    rows_data.append(("data", "TPV ($M)", [round(countries[c]["tpv"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("data", "SoW Cards (%)", [countries[c]["sow_cards"] for c in country_list], "0.00%", None))
    rows_data.append(("data", "TPV Cards ($M)", [round(countries[c]["tpv_cards"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("data", "SoW APMs (%)", [countries[c]["sow_apms"] for c in country_list], "0.00%", None))
    rows_data.append(("data", "TPV APMs ($M)", [round(countries[c]["tpv_apms"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("section", "CARDS BENEFIT"))
    rows_data.append(("data", "Delta Acceptance Rate (%)", [countries[c]["delta_ar"] for c in country_list], "0.00%", None))
    rows_data.append(("data", "Incremental TPV ($M)", [round(countries[c]["incr_tpv"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("data", "Incremental Revenue ($M)", [round(countries[c]["incr_rev"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("data", "Annual Cards TPV ($M)", [round(countries[c]["annual_cards_tpv"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("data", "Local Entity?", [countries[c]["local_entity"] for c in country_list], None, None))
    rows_data.append(("data", "Current MDR", [countries[c]["current_mdr"] for c in country_list], "0.000%", None))
    rows_data.append(("data", "Reduction in MDR", [countries[c]["mdr_reduction"] for c in country_list], "0.000%", None))
    rows_data.append(("data", "New MDR", [countries[c]["new_mdr"] for c in country_list], "0.000%", None))
    rows_data.append(("data", "MDR Reduction Savings ($M)", [round(countries[c]["mdr_savings"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("data", "CARDS BENEFIT ($M)", [round(countries[c]["cards_benefit"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("section", "APMs BENEFIT"))
    rows_data.append(("data", "APMs to Integrate", [countries[c]["apms_to_integrate"] for c in country_list], None, None))
    rows_data.append(("data", "Current APMs (#)", [countries[c]["current_apms_count"] for c in country_list], "#,##0", None))
    rows_data.append(("data", "APMs to Integrate (#)", [countries[c]["apms_count_int"] for c in country_list], "#,##0", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("data", "(%) Incremental Transactions", [countries[c]["pct_incremental_tx"] for c in country_list], "0.00%", None))
    rows_data.append(("data", "APMs Incremental TPV ($M)", [round(countries[c]["apms_incr_tpv"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("data", "APMs Incremental Revenue ($M)", [round(countries[c]["apms_incr_rev"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("data", "(%) TPV from Cards to APMs", [countries[c]["pct_tpv_shift"] for c in country_list], "0.00%", None))
    rows_data.append(("data", "APMs Shifted TPV ($M)", [round(countries[c]["apms_shifted_tpv"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("data", "Cards MDR", [countries[c]["current_mdr"] for c in country_list], "0.000%", None))
    rows_data.append(("data", "APMs MDR Cost", [APMS_MDR_COST] * len(country_list), "0.000%", None))
    rows_data.append(("data", "APM MDR Savings ($M)", [round(countries[c]["apm_mdr_savings"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("data", "APMs BENEFIT ($M)", [round(countries[c]["apms_benefit"], 4) for c in country_list], "#,##0.00", "sum"))
    rows_data.append(("blank",))
    rows_data.append(("data", "TOTAL BENEFIT ($M)", [round(countries[c]["total_benefit"], 2) for c in country_list], "#,##0.00", "sum"))

    # Write header
    write_row(ws, 3, headers, bold=True, fill=total_fill)

    rr = 4
    for row_def in rows_data:
        if row_def[0] == "section":
            section_header(ws, rr, row_def[1], ncols)
            rr += 1
        elif row_def[0] == "blank":
            rr += 1
        else:
            _, label, vals, fmt, total_action = row_def
            row_values = [label] + vals
            if total_action == "sum":
                row_values.append(round(sum(vals), 4))
            else:
                row_values.append(None)
            highlight = "BENEFIT" in label.upper() and "($M)" in label
            write_row(ws, rr, row_values, bold=("BENEFIT" in label.upper()), fill=(total_fill if highlight else None), fmt=fmt)
            rr += 1

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    for i, _ in enumerate(country_list):
        ws.column_dimensions[get_column_letter(3 + i)].width = 22
    ws.column_dimensions[get_column_letter(3 + len(country_list))].width = 18


def build_eng_sheet(region_name, country_list):
    sheet_name = f"BC Devs.Eng {region_name}"
    ws = wb.create_sheet(sheet_name)
    ncols = 3 + len(country_list)
    title_row(ws, f"BC DEVS/ENG — {region_name.upper()}", ncols)

    write_row(ws, 3, ["ENGINEERING INPUTS"] + country_list + ["TOTAL"], bold=True, fill=total_fill)
    rr = 4
    apms_counts = [countries[c]["apms_count_int"] for c in country_list]
    apm_names = [countries[c]["apms_to_integrate"] for c in country_list]
    apm_months = [countries[c]["apm_int_months"] for c in country_list]
    apm_costs = [round(countries[c]["apm_int_cost"], 4) for c in country_list]
    mor_counts = [countries[c]["mor_count"] for c in country_list]
    need_mor = [("Yes" if countries[c]["local_entity"] == "No" else "No") for c in country_list]
    psp_months = [countries[c]["psp_int_months"] for c in country_list]
    psp_costs = [round(countries[c]["psp_int_cost"], 4) for c in country_list]
    total_int_months_list = [countries[c]["total_int_months"] for c in country_list]
    eng_savings_list = [round(countries[c]["eng_savings"], 4) for c in country_list]

    rows = [
        ("APMs to Integrate (#)", apms_counts, "#,##0", True),
        ("APM Names", apm_names, None, False),
        ("Time per APM Integration (months)", [TIME_PER_APM] * len(country_list), "#,##0", False),
        ("APM Integration Months", apm_months, "#,##0", True),
        ("APM Integration Cost ($M)", apm_costs, "#,##0.0000", True),
        ("", None, None, False),
        ("MoR/Processors to Integrate (#)", mor_counts, "#,##0", True),
        ("Need MoR?", need_mor, None, False),
        ("Time per PSP Integration (months)", [TIME_PER_PSP] * len(country_list), "#,##0", False),
        ("PSP Integration Months", psp_months, "#,##0", True),
        ("PSP Integration Cost ($M)", psp_costs, "#,##0.0000", True),
        ("", None, None, False),
        ("Total Integration Months", total_int_months_list, "#,##0", True),
        ("Yuno Integration (months)", [YUNO_INTEGRATION] * len(country_list), "#,##0", False),
        ("TOTAL ENGINEERING SAVINGS ($M)", eng_savings_list, "#,##0.0000", True),
    ]
    for label, vals, fmt, do_sum in rows:
        if label == "":
            rr += 1
            continue
        row_values = [label] + (vals if vals else [None] * len(country_list))
        if do_sum and vals and isinstance(vals[0], (int, float)):
            row_values.append(round(sum(vals), 4))
        else:
            row_values.append(None)
        highlight = "SAVINGS" in label.upper()
        write_row(ws, rr, row_values, bold=highlight, fill=(total_fill if highlight else None), fmt=fmt)
        rr += 1

    rr += 1
    total_months_region = sum(total_int_months_list)
    write_row(ws, rr, ["Total months (all countries)", total_months_region], bold=True, fmt="#,##0"); rr += 1
    write_row(ws, rr, ["Yuno integration (months)", YUNO_INTEGRATION], bold=True, fmt="#,##0"); rr += 1
    write_row(ws, rr, ["Time Saved (months)", total_months_region - YUNO_INTEGRATION], bold=True, fmt="#,##0")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    for i, _ in enumerate(country_list):
        ws.column_dimensions[get_column_letter(3 + i)].width = 28
    ws.column_dimensions[get_column_letter(3 + len(country_list))].width = 18


# Build all regions
for rname, clist in regions.items():
    build_payins_sheet(rname, clist)
    build_eng_sheet(rname, clist)

# ===== Current Payment Methods =====
ws = wb.create_sheet("Current Payment Methods")
title_row(ws, "CURRENT PAYMENT METHODS — COSTCO", 4)
write_row(ws, 3, ["Country", "Current Payment Methods", "Missing / Recommended APMs"], bold=True, fill=total_fill)
rr = 4
for c, d in countries.items():
    write_row(ws, rr, [c, d["current_apms_list"], d["apms_to_integrate"]])
    rr += 1
rr += 1
notes = [
    "1. Costco does not publicly disclose its checkout PSP / acquirer. Elavon is confirmed as the back-end for Costco Payment Processing reseller (B2B sales to members), NOT for Costco.com itself.",
    "2. Single-network exclusive per country: Visa-only US (Citi co-brand), Mastercard-only Canada (CIBC), Mastercard-only UK and Japan, Hyundai-only Korea, Fubon-only Taiwan, Banamex-led Mexico. No public failover/redundancy.",
    "3. Apple Pay quietly removed from Costco.com and app in May 2025 per Frequent Miler. Generated member backlash.",
    "4. Affirm BNPL is US-only and exclusive (May 14 2025 launch, $500 to $17,500, 10-36 percent APR). No BNPL in any other Costco market.",
    "5. Mexico online accepts PayPal but not OXXO/SPEI/Mercado Pago per official help page. Major LATAM APM gap.",
    "6. Australia and Asia checkout details largely unverified (page 403 or membership-gated). Conservative APM assumptions used.",
    "7. Citi Visa exclusive co-brand contract expires 2026 with renewal options through 2029. Public renewal uncertainty per TheStreet 2025.",
]
for n in notes:
    write_row(ws, rr, ["NOTE", n])
    rr += 1

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 75
ws.column_dimensions["D"].width = 60

wb.save(OUT)
print(f"OK: saved to {OUT}")
print(f"Total Benefit: ${tot_benefit:.2f}M")
print(f"  Acceptance Uplift: ${tot_acceptance_uplift:.2f}M")
print(f"  Fee Renegotiation: ${tot_fee_renegotiation:.2f}M")
print(f"  New APM Growth: ${tot_new_apm_growth:.2f}M")
print(f"  Eng Savings: ${tot_eng_savings:.2f}M")
print(f"Time Saved: {time_saved} months ({time_saved_years:.1f} years)")
