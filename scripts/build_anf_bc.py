"""Build BC sheets - Abercrombie & Fitch.xlsx — Yuno Payment Orchestrator BC."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/germantatis/Desktop/GTMCoding/Business Cases/BC sheets - Abercrombie & Fitch.xlsx"

ANNUAL_VISITS = 266_700_000
CONVERSION = 0.06
INDUSTRY = "Retail (Pets, Electronics...)"
TAKE_RATE = 0.30
LOCAL_MDR = 0.018
LOCAL_MDR_REDUCTION = -0.002
LOCAL_MDR_NEW = LOCAL_MDR + LOCAL_MDR_REDUCTION
DELTA_AR_LOCAL = 0.03
APMS_MDR_COST = 0.012
COST_PER_MONTH = 0.0216
TIME_PER_APM = 3
TIME_PER_PSP = 4
YUNO_INTEGRATION = 2
VTC = 0.22
CTS = 0.85

countries = {
    "United States": {
        "region": "NA", "sow": 0.7227, "atv": 130, "local_entity": "Yes",
        "current_apms_list": "Visa, Mastercard, Amex, Discover, Apple Pay, Google Pay, PayPal, Venmo (since 2018), Klarna Pay in 4, Afterpay, Zip; Comenity Bank A&F credit card",
        "current_apms_count": 7,
        "sow_cards": 0.78, "sow_apms": 0.22,
        "apms_to_integrate": "Affirm (gap vs URBN peers), Cash App Pay, Shop Pay",
        "apms_count_int": 3,
        "pct_incremental_tx": 0.03,
        "pct_tpv_shift": 0.015,
        "delta_ar": DELTA_AR_LOCAL,
        "current_mdr": LOCAL_MDR,
        "mdr_reduction": LOCAL_MDR_REDUCTION,
        "new_mdr": LOCAL_MDR_NEW,
        "mor_count": 1,
        "need_mor": "No",
    },
    "Canada": {
        "region": "NA", "sow": 0.0683, "atv": 115, "local_entity": "Yes",
        "current_apms_list": "Cards + Apple Pay (in-app per Hollister iOS listing); Klarna NOT confirmed for Canada; broader Canadian APMs not documented on official pages",
        "current_apms_count": 1,
        "sow_cards": 0.92, "sow_apms": 0.08,
        "apms_to_integrate": "Interac e-Transfer, Interac Online, Klarna (CAD), Google Pay",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.06,
        "pct_tpv_shift": 0.03,
        "delta_ar": DELTA_AR_LOCAL,
        "current_mdr": LOCAL_MDR,
        "mdr_reduction": LOCAL_MDR_REDUCTION,
        "new_mdr": LOCAL_MDR_NEW,
        "mor_count": 1,
        "need_mor": "No",
    },
    "United Kingdom": {
        "region": "EMEA", "sow": 0.1117, "atv": 110, "local_entity": "Yes",
        "current_apms_list": "Visa, Mastercard, Amex, Klarna Pay in 3, PayPal (likely), Apple Pay (in-app); Clearpay/Open Banking NOT confirmed",
        "current_apms_count": 3,
        "sow_cards": 0.85, "sow_apms": 0.15,
        "apms_to_integrate": "Clearpay (direct integration), Open Banking, Google Pay (web), Amazon Pay",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.05,
        "pct_tpv_shift": 0.025,
        "delta_ar": DELTA_AR_LOCAL,
        "current_mdr": LOCAL_MDR,
        "mdr_reduction": LOCAL_MDR_REDUCTION,
        "new_mdr": LOCAL_MDR_NEW,
        "mor_count": 1,
        "need_mor": "No",
    },
    "Germany": {
        "region": "EMEA", "sow": 0.0148, "atv": 115, "local_entity": "Yes",
        "current_apms_list": "Visa, Mastercard, Klarna Pay Later (14-day invoice, DE-original), PayPal (likely), Apple Pay (likely in-app); SEPA/Sofort/Giropay NOT confirmed on official pages",
        "current_apms_count": 3,
        "sow_cards": 0.80, "sow_apms": 0.20,
        "apms_to_integrate": "SEPA Direct Debit, Sofort, Giropay/Wero, Kauf auf Rechnung direct",
        "apms_count_int": 4,
        "pct_incremental_tx": 0.08,
        "pct_tpv_shift": 0.04,
        "delta_ar": DELTA_AR_LOCAL,
        "current_mdr": LOCAL_MDR,
        "mdr_reduction": LOCAL_MDR_REDUCTION,
        "new_mdr": LOCAL_MDR_NEW,
        "mor_count": 1,
        "need_mor": "No",
    },
}

for c, d in countries.items():
    d["visits"] = ANNUAL_VISITS * d["sow"]
    d["txs"] = d["visits"] * CONVERSION
    d["tpv"] = d["txs"] * d["atv"] / 1_000_000
    d["tpv_cards"] = d["tpv"] * d["sow_cards"]
    d["tpv_apms"] = d["tpv"] * d["sow_apms"]
    d["incr_tpv"] = d["tpv_cards"] * d["delta_ar"]
    d["incr_rev"] = d["incr_tpv"] * TAKE_RATE
    d["annual_cards_tpv"] = d["tpv_cards"] + d["incr_tpv"]
    d["mdr_savings"] = d["annual_cards_tpv"] * abs(d["mdr_reduction"])
    d["cards_benefit"] = d["incr_rev"] + d["mdr_savings"]
    d["apms_incr_tpv"] = d["tpv"] * d["pct_incremental_tx"]
    d["apms_incr_rev"] = d["apms_incr_tpv"] * TAKE_RATE
    d["apms_shifted_tpv"] = d["tpv"] * d["pct_tpv_shift"]
    d["apm_mdr_savings"] = d["apms_shifted_tpv"] * (d["current_mdr"] - APMS_MDR_COST)
    d["apms_benefit"] = d["apms_incr_rev"] + d["apm_mdr_savings"]
    d["total_benefit"] = d["cards_benefit"] + d["apms_benefit"]
    d["apm_int_months"] = d["apms_count_int"] * TIME_PER_APM
    d["apm_int_cost"] = d["apm_int_months"] * COST_PER_MONTH
    d["psp_int_months"] = d["mor_count"] * TIME_PER_PSP
    d["psp_int_cost"] = d["psp_int_months"] * COST_PER_MONTH
    d["total_int_months"] = d["apm_int_months"] + d["psp_int_months"]
    d["eng_savings"] = d["apm_int_cost"] + d["psp_int_cost"]

regions = {
    "NA": ["United States", "Canada"],
    "EMEA": ["United Kingdom", "Germany"],
}

tot_acceptance_uplift = sum(d["incr_rev"] for d in countries.values())
tot_fee_renegotiation = sum(d["mdr_savings"] for d in countries.values())
tot_new_apm_growth = sum(d["apms_benefit"] for d in countries.values())
tot_eng_savings = sum(d["eng_savings"] for d in countries.values())
tot_benefit = tot_acceptance_uplift + tot_fee_renegotiation + tot_new_apm_growth + tot_eng_savings
tot_int_months = sum(d["total_int_months"] for d in countries.values())
yuno_total_int_months = YUNO_INTEGRATION * len(regions)
time_saved = tot_int_months - yuno_total_int_months
time_saved_years = time_saved / 12

wb = openpyxl.Workbook()
del wb["Sheet"]
purple_fill = PatternFill(start_color="6B2CB0", end_color="6B2CB0", fill_type="solid")
header_fill = PatternFill(start_color="EFE6FA", end_color="EFE6FA", fill_type="solid")
total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
title_font = Font(bold=True, size=16, color="FFFFFF")
bold_font = Font(bold=True, size=11)
section_font = Font(bold=True, size=11, color="6B2CB0")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def title_row(ws, txt, ncols):
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=2, value=txt)
    cell.font = title_font; cell.fill = purple_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

def write_row(ws, row, vals, bold=False, fill=None, fmt=None, indent_col=2):
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=indent_col + i, value=v)
        if bold: cell.font = bold_font
        if fill: cell.fill = fill
        if fmt and isinstance(v, (int, float)) and i > 0:
            cell.number_format = fmt
        cell.border = border

def section_header(ws, row, txt, ncols):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=2, value=txt)
    c.font = section_font; c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="center")

# RESUMEN
ws = wb.create_sheet("RESUMEN")
title_row(ws, "BUSINESS CASE — ABERCROMBIE & FITCH", 8)
r = 3
write_row(ws, r, ["Company", "Abercrombie & Fitch Co."], bold=True); r += 1
write_row(ws, r, ["Industry", INDUSTRY], bold=True); r += 1
write_row(ws, r, ["Annual Total Visits", ANNUAL_VISITS], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Annual Revenue ($M)", 4950], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Take Rate", TAKE_RATE], bold=True, fmt="0.00%"); r += 2

section_header(ws, r, "TOTAL BENEFIT SUMMARY", 8); r += 1
write_row(ws, r, ["Category", "Value ($M)", "Lower Range ($M)", "Upper Range ($M)"], bold=True, fill=total_fill); r += 1
def lo_hi(v): return [round(v, 2), round(v * 0.85, 2), round(v * 1.15, 2)]
write_row(ws, r, ["TPV Acceptance Rate Uplift"] + lo_hi(tot_acceptance_uplift), fmt="#,##0.00"); r += 1
write_row(ws, r, ["Fee Renegotiation (MDR Savings)"] + lo_hi(tot_fee_renegotiation), fmt="#,##0.00"); r += 1
write_row(ws, r, ["New APM TPV Growth"] + lo_hi(tot_new_apm_growth), fmt="#,##0.00"); r += 1
write_row(ws, r, ["Engineering Cost Savings"] + lo_hi(tot_eng_savings), fmt="#,##0.00"); r += 1
write_row(ws, r, ["TOTAL BENEFIT"] + lo_hi(tot_benefit), bold=True, fill=total_fill, fmt="#,##0.00"); r += 2

write_row(ws, r, ["Total integration months (without Yuno)", tot_int_months], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Yuno integration time (months, per region x2)", yuno_total_int_months], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Time Saved (months)", time_saved], bold=True, fmt="#,##0"); r += 1
write_row(ws, r, ["Time Saved (years)", round(time_saved_years, 1)], bold=True, fmt="0.0"); r += 2

problems = [
    "PSP/acquirer for abercrombie.com, hollisterco.com and gillyhicks.com is not publicly disclosed; no public failover, smart routing, or multi-acquirer evidence above the card stack",
    "Cross-border international shipments rely on ESW (eShopWorld) as merchant of record, surrendering margin and limiting direct control over local APMs and routing in markets without a local entity",
    "Germany APM stack incomplete: Klarna Pay Later is in place, but SEPA Direct Debit, Sofort, Giropay and Wero (where PayPal #1 and Kauf auf Rechnung are 82%-abandonment APMs) are not confirmed on official pages",
    "Canada has documented gap: no Interac e-Transfer or Interac Online on official help pages despite Interac being the dominant local debit rail",
    "New ERP rollout flagged as Q1 2026 risk factor by CFO Robert Ball; vendor change window opens broader payment-stack revaluation under same transformation umbrella",
]
for i, p in enumerate(problems, 1):
    write_row(ws, r, [None, None, None, None, None, f"PROBLEM {i}", p], indent_col=2); r += 1

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 4
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 110

def build_payins_sheet(region_name, country_list):
    ws = wb.create_sheet(f"BC Pay-ins {region_name}")
    ncols = 3 + len(country_list)
    title_row(ws, f"BC PAY-INS — {region_name.upper()}", ncols)
    write_row(ws, 3, ["SDR INPUTS"] + country_list + ["TOTAL"], bold=True, fill=total_fill)
    rows_data = [
        ("section", "TRAFFIC & TPV"),
        ("data", "Web Traffic (SoW)", [countries[c]["sow"] for c in country_list], "0.00%", None),
        ("data", "Web Traffic Visits", [int(countries[c]["visits"]) for c in country_list], "#,##0", None),
        ("data", "Conversion", [CONVERSION] * len(country_list), "0.00%", None),
        ("data", "View to Checkout (VtC)", [VTC] * len(country_list), "0.00%", None),
        ("data", "Checkout to Success (CtS)", [CTS] * len(country_list), "0.00%", None),
        ("data", "Successful Transactions/Year", [int(countries[c]["txs"]) for c in country_list], "#,##0", "sum"),
        ("data", "ATV ($)", [countries[c]["atv"] for c in country_list], "$#,##0", None),
        ("data", "TPV ($M)", [round(countries[c]["tpv"], 4) for c in country_list], "#,##0.00", "sum"),
        ("data", "SoW Cards (%)", [countries[c]["sow_cards"] for c in country_list], "0.00%", None),
        ("data", "TPV Cards ($M)", [round(countries[c]["tpv_cards"], 4) for c in country_list], "#,##0.00", "sum"),
        ("data", "SoW APMs (%)", [countries[c]["sow_apms"] for c in country_list], "0.00%", None),
        ("data", "TPV APMs ($M)", [round(countries[c]["tpv_apms"], 4) for c in country_list], "#,##0.00", "sum"),
        ("blank",),
        ("section", "CARDS BENEFIT"),
        ("data", "Delta Acceptance Rate (%)", [countries[c]["delta_ar"] for c in country_list], "0.00%", None),
        ("data", "Incremental TPV ($M)", [round(countries[c]["incr_tpv"], 4) for c in country_list], "#,##0.00", "sum"),
        ("data", "Incremental Revenue ($M)", [round(countries[c]["incr_rev"], 4) for c in country_list], "#,##0.00", "sum"),
        ("blank",),
        ("data", "Annual Cards TPV ($M)", [round(countries[c]["annual_cards_tpv"], 4) for c in country_list], "#,##0.00", "sum"),
        ("data", "Local Entity?", [countries[c]["local_entity"] for c in country_list], None, None),
        ("data", "Current MDR", [countries[c]["current_mdr"] for c in country_list], "0.000%", None),
        ("data", "Reduction in MDR", [countries[c]["mdr_reduction"] for c in country_list], "0.000%", None),
        ("data", "New MDR", [countries[c]["new_mdr"] for c in country_list], "0.000%", None),
        ("data", "MDR Reduction Savings ($M)", [round(countries[c]["mdr_savings"], 4) for c in country_list], "#,##0.00", "sum"),
        ("blank",),
        ("data", "CARDS BENEFIT ($M)", [round(countries[c]["cards_benefit"], 4) for c in country_list], "#,##0.00", "sum"),
        ("blank",),
        ("section", "APMs BENEFIT"),
        ("data", "APMs to Integrate", [countries[c]["apms_to_integrate"] for c in country_list], None, None),
        ("data", "Current APMs (#)", [countries[c]["current_apms_count"] for c in country_list], "#,##0", None),
        ("data", "APMs to Integrate (#)", [countries[c]["apms_count_int"] for c in country_list], "#,##0", "sum"),
        ("blank",),
        ("data", "(%) Incremental Transactions", [countries[c]["pct_incremental_tx"] for c in country_list], "0.00%", None),
        ("data", "APMs Incremental TPV ($M)", [round(countries[c]["apms_incr_tpv"], 4) for c in country_list], "#,##0.00", "sum"),
        ("data", "APMs Incremental Revenue ($M)", [round(countries[c]["apms_incr_rev"], 4) for c in country_list], "#,##0.00", "sum"),
        ("blank",),
        ("data", "(%) TPV from Cards to APMs", [countries[c]["pct_tpv_shift"] for c in country_list], "0.00%", None),
        ("data", "APMs Shifted TPV ($M)", [round(countries[c]["apms_shifted_tpv"], 4) for c in country_list], "#,##0.00", "sum"),
        ("blank",),
        ("data", "Cards MDR", [countries[c]["current_mdr"] for c in country_list], "0.000%", None),
        ("data", "APMs MDR Cost", [APMS_MDR_COST] * len(country_list), "0.000%", None),
        ("data", "APM MDR Savings ($M)", [round(countries[c]["apm_mdr_savings"], 4) for c in country_list], "#,##0.00", "sum"),
        ("blank",),
        ("data", "APMs BENEFIT ($M)", [round(countries[c]["apms_benefit"], 4) for c in country_list], "#,##0.00", "sum"),
        ("blank",),
        ("data", "TOTAL BENEFIT ($M)", [round(countries[c]["total_benefit"], 2) for c in country_list], "#,##0.00", "sum"),
    ]
    rr = 4
    for row_def in rows_data:
        if row_def[0] == "section":
            section_header(ws, rr, row_def[1], ncols); rr += 1
        elif row_def[0] == "blank":
            rr += 1
        else:
            _, label, vals, fmt, total_action = row_def
            row_values = [label] + vals
            if total_action == "sum" and all(isinstance(v, (int, float)) for v in vals):
                row_values.append(round(sum(vals), 4))
            else:
                row_values.append(None)
            highlight = "BENEFIT" in label.upper() and "($M)" in label
            write_row(ws, rr, row_values, bold=("BENEFIT" in label.upper()),
                      fill=(total_fill if highlight else None), fmt=fmt)
            rr += 1
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    for i, _ in enumerate(country_list):
        ws.column_dimensions[get_column_letter(3 + i)].width = 28
    ws.column_dimensions[get_column_letter(3 + len(country_list))].width = 18

def build_eng_sheet(region_name, country_list):
    ws = wb.create_sheet(f"BC Devs.Eng {region_name}")
    ncols = 3 + len(country_list)
    title_row(ws, f"BC DEVS/ENG — {region_name.upper()}", ncols)
    write_row(ws, 3, ["ENGINEERING INPUTS"] + country_list + ["TOTAL"], bold=True, fill=total_fill)
    rr = 4
    apms_counts = [countries[c]["apms_count_int"] for c in country_list]
    apm_names = [countries[c]["apms_to_integrate"] for c in country_list]
    apm_months = [countries[c]["apm_int_months"] for c in country_list]
    apm_costs = [round(countries[c]["apm_int_cost"], 4) for c in country_list]
    mor_counts = [countries[c]["mor_count"] for c in country_list]
    need_mor = [countries[c]["need_mor"] for c in country_list]
    psp_months = [countries[c]["psp_int_months"] for c in country_list]
    psp_costs = [round(countries[c]["psp_int_cost"], 4) for c in country_list]
    total_int_months_list = [countries[c]["total_int_months"] for c in country_list]
    eng_savings_list = [round(countries[c]["eng_savings"], 4) for c in country_list]
    rows = [
        ("APMs to Integrate (#)", apms_counts, "#,##0", True),
        ("APM Names", apm_names, None, False),
        ("Time per APM Integration (months)", [TIME_PER_APM] * len(country_list), "#,##0", False),
        ("APM Integration Months", apm_months, "#,##0", True),
        ("APM Integration Cost ($M)", apm_costs, "#,##0.0000", True),
        ("", None, None, False),
        ("MoR/Processors to Integrate (#)", mor_counts, "#,##0", True),
        ("Need MoR?", need_mor, None, False),
        ("Time per PSP Integration (months)", [TIME_PER_PSP] * len(country_list), "#,##0", False),
        ("PSP Integration Months", psp_months, "#,##0", True),
        ("PSP Integration Cost ($M)", psp_costs, "#,##0.0000", True),
        ("", None, None, False),
        ("Total Integration Months", total_int_months_list, "#,##0", True),
        ("Yuno Integration (months)", [YUNO_INTEGRATION] * len(country_list), "#,##0", False),
        ("TOTAL ENGINEERING SAVINGS ($M)", eng_savings_list, "#,##0.0000", True),
    ]
    for label, vals, fmt, do_sum in rows:
        if label == "":
            rr += 1; continue
        row_values = [label] + (vals if vals else [None] * len(country_list))
        if do_sum and vals and isinstance(vals[0], (int, float)):
            row_values.append(round(sum(vals), 4))
        else:
            row_values.append(None)
        highlight = "SAVINGS" in label.upper()
        write_row(ws, rr, row_values, bold=highlight, fill=(total_fill if highlight else None), fmt=fmt)
        rr += 1
    rr += 1
    total_months_region = sum(total_int_months_list)
    write_row(ws, rr, ["Total months (all countries)", total_months_region], bold=True, fmt="#,##0"); rr += 1
    write_row(ws, rr, ["Yuno integration (months)", YUNO_INTEGRATION], bold=True, fmt="#,##0"); rr += 1
    write_row(ws, rr, ["Time Saved (months)", total_months_region - YUNO_INTEGRATION], bold=True, fmt="#,##0")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    for i, _ in enumerate(country_list):
        ws.column_dimensions[get_column_letter(3 + i)].width = 28
    ws.column_dimensions[get_column_letter(3 + len(country_list))].width = 18

for rname, clist in regions.items():
    build_payins_sheet(rname, clist)
    build_eng_sheet(rname, clist)

ws = wb.create_sheet("Current Payment Methods")
title_row(ws, "CURRENT PAYMENT METHODS — ABERCROMBIE & FITCH", 4)
write_row(ws, 3, ["Country", "Current Payment Methods", "Missing / Recommended APMs"], bold=True, fill=total_fill)
rr = 4
for c, d in countries.items():
    write_row(ws, rr, [c, d["current_apms_list"], d["apms_to_integrate"]]); rr += 1
rr += 1
notes = [
    "1. PSP/acquirer for abercrombie.com, hollisterco.com, gillyhicks.com is NOT publicly disclosed. Likely Cybersource or Adyen given Salesforce Commerce Cloud (Demandware) + ESW combo, but unconfirmed.",
    "2. International cross-border ships via ESW (eShopWorld) as merchant of record; cardholder statement shows ESW, not A&F.",
    "3. New ERP system rolled out FY2025; CFO Robert Ball cited ERP transition as Q1 2026 risk factor in earnings call.",
    "4. BNPL: Klarna is primary partner (Germany since pre-2019, US Pay in 4 since Oct 2019, UK Pay in 3). Plus Afterpay and Zip in US (per third-party aggregators). Affirm gap vs URBN peer set.",
    "5. Venmo: A&F was first specialty retailer to offer Venmo (Aug 2018, in-app for A&F and Hollister).",
    "6. Refund-to-original-payment-method failures dominate BBB/Trustpilot complaints. Forced e-gift card refunds and 3-week refund delays (A&F worse than Hollister despite shared facility).",
    "7. Comenity Bank A&F private label card: account portal active May 2026; one source claims card discontinued May 2020 (conflicting). APR 26.49% per CFPB filing.",
    "8. Canada research gap: no Interac e-Transfer or Interac Online on official help pages despite Interac being dominant local rail.",
    "9. Germany research gap: SEPA Direct Debit, Sofort, Giropay status NOT confirmed on official pages (PayPal is #1 in Germany ecom; Kauf auf Rechnung drives 82% abandonment if missing).",
    "10. myAbercrombie loyalty: 70-80% of A&F customers enrolled (Talon.One); VIP tier $500+ annual spend.",
]
for n in notes:
    write_row(ws, rr, ["NOTE", n]); rr += 1
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 75
ws.column_dimensions["D"].width = 60

wb.save(OUT)
print(f"OK: saved to {OUT}")
print(f"Total Benefit: ${tot_benefit:.2f}M")
print(f"  Acceptance Uplift: ${tot_acceptance_uplift:.2f}M")
print(f"  Fee Renegotiation: ${tot_fee_renegotiation:.2f}M")
print(f"  New APM Growth: ${tot_new_apm_growth:.2f}M")
print(f"  Eng Savings: ${tot_eng_savings:.2f}M")
print(f"Time Saved: {time_saved} months ({time_saved_years:.1f} years)")
