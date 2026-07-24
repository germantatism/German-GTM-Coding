"""Build BC sheets - Cettire.xlsx mirroring the Nordstrom/Abercrombie template."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- INPUTS ----------
ANNUAL_VISITS = 39_910_000
CONVERSION = 0.06
VTC = 0.10
CTS = 0.60
TAKE_RATE = 0.22  # Cettire FY25 delivered margin 16.1% blended up to reflect gross-revenue take vs sales
APM_MDR_COST = 0.012
MONTHS_COST = 21_600  # blended team cost/month per integration (matches Nordstrom/Abercrombie)
TIME_APM = 3
TIME_PSP = 4
YUNO_MONTHS = 2  # per region

# Industry benchmark — Luxury goods/Apparel
LOCAL_MDR = 0.018
CB_MDR = 0.042
DELTA_LOCAL = 0.01
DELTA_CB = 0.03

COUNTRIES = {
    # NA
    "United States":  {"region": "NA",   "sow": 0.3292, "atv": 850,  "entity": "No",  "current_apms": 7, "to_integrate": ["PayPal", "Venmo", "Cash App Pay"],                                     "sow_cards": 0.78, "pct_inc": 0.06, "pct_shift": 0.025},
    "Canada":         {"region": "NA",   "sow": 0.0672, "atv": 880,  "entity": "No",  "current_apms": 4, "to_integrate": ["Interac e-Transfer", "Interac Online", "PayPal"],                       "sow_cards": 0.88, "pct_inc": 0.08, "pct_shift": 0.04},
    "Mexico":         {"region": "NA",   "sow": 0.0125, "atv": 620,  "entity": "No",  "current_apms": 2, "to_integrate": ["OXXO Pay", "SPEI", "Mercado Pago", "MSI Installments"],               "sow_cards": 0.94, "pct_inc": 0.10, "pct_shift": 0.05},
    # EMEA
    "United Kingdom": {"region": "EMEA", "sow": 0.0804, "atv": 900,  "entity": "Yes", "current_apms": 5, "to_integrate": ["PayPal", "Open Banking"],                                              "sow_cards": 0.82, "pct_inc": 0.06, "pct_shift": 0.025},
    "Germany":        {"region": "EMEA", "sow": 0.0295, "atv": 880,  "entity": "No",  "current_apms": 4, "to_integrate": ["PayPal", "SEPA Direct Debit", "Sofort", "Giropay"],                  "sow_cards": 0.88, "pct_inc": 0.10, "pct_shift": 0.05},
    "Netherlands":    {"region": "EMEA", "sow": 0.0147, "atv": 870,  "entity": "No",  "current_apms": 4, "to_integrate": ["iDEAL", "PayPal"],                                                     "sow_cards": 0.92, "pct_inc": 0.12, "pct_shift": 0.05},
    "Saudi Arabia":   {"region": "EMEA", "sow": 0.0108, "atv": 1450, "entity": "No",  "current_apms": 2, "to_integrate": ["Mada", "STC Pay", "Tabby", "Tamara"],                                  "sow_cards": 0.92, "pct_inc": 0.10, "pct_shift": 0.04},
    "UAE":            {"region": "EMEA", "sow": 0.0106, "atv": 1500, "entity": "No",  "current_apms": 2, "to_integrate": ["Tabby", "Tamara", "Careem Pay"],                                       "sow_cards": 0.90, "pct_inc": 0.08, "pct_shift": 0.03},
    # APAC
    "South Korea":    {"region": "APAC", "sow": 0.1153, "atv": 1050, "entity": "No",  "current_apms": 2, "to_integrate": ["KakaoPay", "Naver Pay", "Toss", "Samsung Pay"],                       "sow_cards": 0.94, "pct_inc": 0.10, "pct_shift": 0.05},
    "Japan":          {"region": "APAC", "sow": 0.1121, "atv": 780,  "entity": "No",  "current_apms": 2, "to_integrate": ["JCB Local Acquiring", "Konbini", "PayPay", "Rakuten Pay"],            "sow_cards": 0.93, "pct_inc": 0.08, "pct_shift": 0.04},
    "Australia":      {"region": "APAC", "sow": 0.0698, "atv": 920,  "entity": "Yes", "current_apms": 6, "to_integrate": ["PayPal", "NPP/Osko", "BPAY"],                                          "sow_cards": 0.80, "pct_inc": 0.04, "pct_shift": 0.02},
    "Taiwan":         {"region": "APAC", "sow": 0.0114, "atv": 900,  "entity": "No",  "current_apms": 2, "to_integrate": ["LINE Pay", "JKO Pay", "JCB Local Acquiring"],                          "sow_cards": 0.92, "pct_inc": 0.06, "pct_shift": 0.025},
    "Hong Kong":      {"region": "APAC", "sow": 0.0112, "atv": 1250, "entity": "No",  "current_apms": 2, "to_integrate": ["Octopus", "AlipayHK", "FPS", "PayMe"],                                  "sow_cards": 0.90, "pct_inc": 0.06, "pct_shift": 0.03},
}

CURRENT_PMS = {
    "United States":  "Visa, Mastercard, Amex; Apple Pay, Google Pay, Shop Pay; Afterpay, Klarna, Zip, Affirm",
    "Canada":         "Visa, Mastercard, Amex; Apple Pay, Google Pay, Shop Pay; Klarna",
    "Mexico":         "Visa, Mastercard, Amex; Apple Pay, Google Pay",
    "United Kingdom": "Visa, Mastercard, Amex; Apple Pay, Google Pay, Shop Pay; Clearpay (Afterpay UK), Klarna",
    "Germany":        "Visa, Mastercard, Amex; Apple Pay, Google Pay, Shop Pay; Klarna",
    "Netherlands":    "Visa, Mastercard, Amex; Apple Pay, Google Pay, Shop Pay; Klarna",
    "Saudi Arabia":   "Visa, Mastercard, Amex; Apple Pay, Google Pay",
    "UAE":            "Visa, Mastercard, Amex; Apple Pay, Google Pay",
    "South Korea":    "Visa, Mastercard, Amex; Apple Pay, Google Pay",
    "Japan":          "Visa, Mastercard, Amex; Apple Pay, Google Pay",
    "Australia":      "Visa, Mastercard, Amex; Apple Pay, Google Pay, Shop Pay; Afterpay, Klarna, Zip",
    "Taiwan":         "Visa, Mastercard, Amex; Apple Pay, Google Pay",
    "Hong Kong":      "Visa, Mastercard, Amex; Apple Pay, Google Pay",
}


def compute(c):
    visits = ANNUAL_VISITS * c["sow"]
    txns = visits * CONVERSION
    tpv = txns * c["atv"]
    tpv_cards = tpv * c["sow_cards"]
    sow_apms = 1 - c["sow_cards"]
    tpv_apms = tpv * sow_apms
    delta = DELTA_LOCAL if c["entity"] == "Yes" else DELTA_CB
    inc_tpv = tpv_cards * delta
    inc_rev = inc_tpv * TAKE_RATE
    annual_cards_tpv = tpv_cards + inc_tpv
    current_mdr = LOCAL_MDR if c["entity"] == "Yes" else CB_MDR
    mdr_reduction = -0.002 if c["entity"] == "Yes" else -0.01
    new_mdr = current_mdr + mdr_reduction
    mdr_savings = annual_cards_tpv * abs(mdr_reduction)
    cards_benefit = inc_rev + mdr_savings

    apm_inc_tpv = tpv * c["pct_inc"]
    apm_inc_rev = apm_inc_tpv * TAKE_RATE
    apm_shifted_tpv = tpv * c["pct_shift"]
    apm_mdr_savings = apm_shifted_tpv * (current_mdr - APM_MDR_COST)
    apms_benefit = apm_inc_rev + apm_mdr_savings

    total_benefit = cards_benefit + apms_benefit

    return {
        "visits": visits, "txns": txns, "tpv": tpv,
        "tpv_cards": tpv_cards, "sow_apms": sow_apms, "tpv_apms": tpv_apms,
        "delta": delta, "inc_tpv": inc_tpv, "inc_rev": inc_rev,
        "annual_cards_tpv": annual_cards_tpv,
        "current_mdr": current_mdr, "mdr_reduction": mdr_reduction, "new_mdr": new_mdr,
        "mdr_savings": mdr_savings, "cards_benefit": cards_benefit,
        "apm_inc_tpv": apm_inc_tpv, "apm_inc_rev": apm_inc_rev,
        "apm_shifted_tpv": apm_shifted_tpv, "apm_mdr_savings": apm_mdr_savings,
        "apms_benefit": apms_benefit, "total_benefit": total_benefit,
    }


# Pre-compute
RESULTS = {name: compute(c) for name, c in COUNTRIES.items()}

# Regional totals
REGIONS = {"NA": [], "EMEA": [], "APAC": []}
for name, c in COUNTRIES.items():
    REGIONS[c["region"]].append(name)

# Global totals for RESUMEN (in $M)
def millions(x): return x / 1_000_000

TOTAL_INC_REV = sum(r["inc_rev"] for r in RESULTS.values())
TOTAL_MDR_SAVINGS = sum(r["mdr_savings"] for r in RESULTS.values())
TOTAL_APM_INC_REV = sum(r["apm_inc_rev"] for r in RESULTS.values())
TOTAL_APM_MDR_SAVINGS = sum(r["apm_mdr_savings"] for r in RESULTS.values())
TOTAL_APM_BENEFIT = TOTAL_APM_INC_REV + TOTAL_APM_MDR_SAVINGS

# Engineering savings
def eng_for_country(name):
    c = COUNTRIES[name]
    n_apms = len(c["to_integrate"])
    apm_months = n_apms * TIME_APM
    apm_cost = apm_months * MONTHS_COST
    n_mors = 2 if c["entity"] == "No" else 1
    need_mor = "Yes" if c["entity"] == "No" else "No"
    psp_months = n_mors * TIME_PSP
    psp_cost = psp_months * MONTHS_COST
    total_months = apm_months + psp_months
    total_savings = apm_cost + psp_cost
    return {
        "n_apms": n_apms, "apm_months": apm_months, "apm_cost": apm_cost,
        "n_mors": n_mors, "need_mor": need_mor, "psp_months": psp_months, "psp_cost": psp_cost,
        "total_months": total_months, "total_savings": total_savings,
    }

ENG_RESULTS = {name: eng_for_country(name) for name in COUNTRIES}
TOTAL_ENG_SAVINGS = sum(e["total_savings"] for e in ENG_RESULTS.values())
TOTAL_INTEGRATION_MONTHS = sum(e["total_months"] for e in ENG_RESULTS.values())
YUNO_TOTAL_MONTHS = YUNO_MONTHS * 3  # 3 regions
TIME_SAVED_MONTHS = TOTAL_INTEGRATION_MONTHS - YUNO_TOTAL_MONTHS

TOTAL_BENEFIT_M = millions(TOTAL_INC_REV + TOTAL_MDR_SAVINGS + TOTAL_APM_BENEFIT + TOTAL_ENG_SAVINGS)

# ---------- BUILD WORKBOOK ----------
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
SECTION_FONT = Font(bold=True, color="111827", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="FEF3C7")
TOTAL_FONT = Font(bold=True, size=11)
LABEL_FONT = Font(size=10)
BORDER = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
                top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))

# ===== RESUMEN =====
ws = wb.create_sheet("RESUMEN")
ws["B1"] = "BUSINESS CASE — CETTIRE"
ws["B1"].font = Font(bold=True, size=16, color="FFFFFF")
ws["B1"].fill = HDR_FILL
ws.merge_cells("B1:H1")

ws["B3"] = "Company";              ws["C3"] = "Cettire Limited (ASX: CTT)"
ws["B4"] = "Industry";             ws["C4"] = "Luxury goods/Apparel (cross-border drop-ship)"
ws["B5"] = "Annual Total Visits";  ws["C5"] = ANNUAL_VISITS
ws["B6"] = "Annual Revenue ($M)";  ws["C6"] = 485   # FY25 USD est.
ws["B7"] = "Take Rate";            ws["C7"] = TAKE_RATE

ws["B9"] = "TOTAL BENEFIT SUMMARY"
ws["B9"].font = SECTION_FONT; ws["B9"].fill = SECTION_FILL
ws.merge_cells("B9:E9")

ws["B10"] = "Category"; ws["C10"] = "Value ($M)"; ws["D10"] = "Lower Range ($M)"; ws["E10"] = "Upper Range ($M)"
for c in "BCDE":
    ws[f"{c}10"].font = HDR_FONT; ws[f"{c}10"].fill = HDR_FILL

acc_uplift = millions(TOTAL_INC_REV)
fee_reno = millions(TOTAL_MDR_SAVINGS)
apm_growth = millions(TOTAL_APM_BENEFIT)
eng_savings_m = millions(TOTAL_ENG_SAVINGS)
total = acc_uplift + fee_reno + apm_growth + eng_savings_m

rows = [
    ("TPV Acceptance Rate Uplift",     acc_uplift),
    ("Fee Renegotiation (MDR Savings)", fee_reno),
    ("New APM TPV Growth",              apm_growth),
    ("Engineering Cost Savings",        eng_savings_m),
    ("TOTAL BENEFIT",                   total),
]
for i, (label, val) in enumerate(rows, start=11):
    ws[f"B{i}"] = label
    ws[f"C{i}"] = round(val, 2)
    ws[f"D{i}"] = round(val * 0.85, 2)
    ws[f"E{i}"] = round(val * 1.15, 2)
    if label == "TOTAL BENEFIT":
        for c in "BCDE":
            ws[f"{c}{i}"].font = TOTAL_FONT; ws[f"{c}{i}"].fill = TOTAL_FILL

ws["B17"] = "Total integration months (without Yuno)"; ws["C17"] = TOTAL_INTEGRATION_MONTHS
ws["B18"] = "Yuno integration time (months, per region x3)"; ws["C18"] = YUNO_TOTAL_MONTHS
ws["B19"] = "Time Saved (months)";                     ws["C19"] = TIME_SAVED_MONTHS
ws["B20"] = "Time Saved (years)";                      ws["C20"] = round(TIME_SAVED_MONTHS / 12, 1)

problems = [
    ("PROBLEM 1", "Single card-acquiring rail (Shopify Payments / Stripe) across all 13 high-traffic markets; no public failover, smart routing, or multi-acquirer evidence"),
    ("PROBLEM 2", "Cross-border merchant of record (Cettire Pty Ltd, AU) with no local entity in 11 of 13 markets (only AU parent and UK Co. #10355981 confirmed); pushes ~95% of TPV through cross-border MDR (~4.2%) vs. local (~1.8%)"),
    ("PROBLEM 3", "Zero local APM coverage anywhere: missing iDEAL (NL 70% of e-com), Mada (SA 93%), KakaoPay/Naver (KR ~25%), Konbini/PayPay (JP ~20%), Tabby/Tamara (UAE/SA ~8%), Interac (CA ~20%), OXXO/SPEI (MX ~25%); no PayPal globally"),
    ("PROBLEM 4", "Documented refund and dispute friction at scale: Trustpilot 1.5/5, ProductReview.com.au 1.5/5 (196 reviews), PissedConsumer 222 one-star reviews; AFR duty-remittance exposé Mar 2024 (-27% stock day-one); chargebacks repeatedly cited as customers' only working resolution path"),
    ("PROBLEM 5", "US de minimis repeal (Aug 2025) hit ~40% of US gross revenue at the $800 exemption; FY25 net loss + delivered margin collapse 20.9%→16.1% force localization pivot into KR/JP/MENA/CN that the current payment stack cannot serve"),
]
for i, (k, v) in enumerate(problems, start=22):
    ws[f"G{i}"] = k; ws[f"H{i}"] = v
    ws[f"G{i}"].font = Font(bold=True)
    ws[f"H{i}"].alignment = Alignment(wrap_text=True, vertical="top")

# Column widths
for col, w in [("B", 38), ("C", 22), ("D", 20), ("E", 20), ("G", 12), ("H", 110)]:
    ws.column_dimensions[col].width = w
for r in range(22, 27):
    ws.row_dimensions[r].height = 60


# ===== PAY-INS SHEETS =====
def build_payins(region_name, countries):
    ws = wb.create_sheet(f"BC Pay-ins {region_name}")
    ws[f"B1"] = f"BC PAY-INS — {region_name}"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF"); ws["B1"].fill = HDR_FILL
    last_col = get_column_letter(2 + len(countries) + 1)
    ws.merge_cells(f"B1:{last_col}1")

    ws["B3"] = "SDR INPUTS"
    for i, name in enumerate(countries):
        ws.cell(row=3, column=3+i, value=name)
    ws.cell(row=3, column=3+len(countries), value="TOTAL")
    for c in range(2, 3+len(countries)+1):
        ws.cell(row=3, column=c).font = HDR_FONT; ws.cell(row=3, column=c).fill = HDR_FILL

    def row_section(r, label):
        ws.cell(row=r, column=2, value=label).font = SECTION_FONT
        ws.cell(row=r, column=2).fill = SECTION_FILL

    def row_data(r, label, values, totalize=True, fmt=None, total_override=None):
        ws.cell(row=r, column=2, value=label)
        total = 0
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=3+i, value=v)
            if isinstance(v, (int, float)) and fmt:
                cell.number_format = fmt
            if isinstance(v, (int, float)) and totalize:
                total += v
        if totalize and isinstance(values[0], (int, float)):
            c = ws.cell(row=r, column=3+len(countries), value=round(total_override if total_override is not None else total, 4))
            if fmt: c.number_format = fmt
            c.font = Font(bold=True)

    r = 4
    row_section(r, "TRAFFIC & TPV"); r += 1
    row_data(r, "Web Traffic (SoW)", [round(COUNTRIES[n]["sow"], 4) for n in countries], totalize=False);
    for i,n in enumerate(countries):
        ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Web Traffic Visits", [int(RESULTS[n]["visits"]) for n in countries], fmt="#,##0"); r += 1
    row_data(r, "Conversion", [CONVERSION]*len(countries), totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "View to Checkout (VtC)", [VTC]*len(countries), totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Checkout to Success (CtS)", [CTS]*len(countries), totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Successful Transactions/Year", [int(RESULTS[n]["txns"]) for n in countries], fmt="#,##0"); r += 1
    row_data(r, "ATV ($)", [COUNTRIES[n]["atv"] for n in countries], totalize=False); r += 1
    row_data(r, "TPV ($M)", [round(millions(RESULTS[n]["tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "SoW Cards (%)", [COUNTRIES[n]["sow_cards"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "TPV Cards ($M)", [round(millions(RESULTS[n]["tpv_cards"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "SoW APMs (%)", [round(RESULTS[n]["sow_apms"], 4) for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "TPV APMs ($M)", [round(millions(RESULTS[n]["tpv_apms"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_section(r, "CARDS BENEFIT"); r += 1
    row_data(r, "Delta Acceptance Rate (%)", [RESULTS[n]["delta"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Incremental TPV ($M)", [round(millions(RESULTS[n]["inc_tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "Incremental Revenue ($M)", [round(millions(RESULTS[n]["inc_rev"]), 4) for n in countries], fmt="#,##0.00"); r += 2
    row_data(r, "Annual Cards TPV ($M)", [round(millions(RESULTS[n]["annual_cards_tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "Local Entity?", [COUNTRIES[n]["entity"] for n in countries], totalize=False); r += 1
    row_data(r, "Current MDR", [RESULTS[n]["current_mdr"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Reduction in MDR", [RESULTS[n]["mdr_reduction"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "New MDR", [RESULTS[n]["new_mdr"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "MDR Reduction Savings ($M)", [round(millions(RESULTS[n]["mdr_savings"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_data(r, "CARDS BENEFIT ($M)", [round(millions(RESULTS[n]["cards_benefit"]), 4) for n in countries], fmt="#,##0.00")
    for c_idx in range(2, 3+len(countries)+1):
        ws.cell(row=r, column=c_idx).font = TOTAL_FONT; ws.cell(row=r, column=c_idx).fill = TOTAL_FILL
    r += 2

    row_section(r, "APMs BENEFIT"); r += 1
    row_data(r, "APMs to Integrate", [", ".join(COUNTRIES[n]["to_integrate"]) for n in countries], totalize=False)
    for i in range(len(countries)):
        ws.cell(row=r, column=3+i).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 45
    r += 1
    row_data(r, "Current APMs (#)", [COUNTRIES[n]["current_apms"] for n in countries], fmt="0"); r += 1
    row_data(r, "APMs to Integrate (#)", [len(COUNTRIES[n]["to_integrate"]) for n in countries], fmt="0"); r += 2

    row_data(r, "(%) Incremental Transactions", [COUNTRIES[n]["pct_inc"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "APMs Incremental TPV ($M)", [round(millions(RESULTS[n]["apm_inc_tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "APMs Incremental Revenue ($M)", [round(millions(RESULTS[n]["apm_inc_rev"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_data(r, "(%) TPV from Cards to APMs", [COUNTRIES[n]["pct_shift"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "APMs Shifted TPV ($M)", [round(millions(RESULTS[n]["apm_shifted_tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_data(r, "Cards MDR", [RESULTS[n]["current_mdr"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "APMs MDR Cost", [APM_MDR_COST]*len(countries), totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "APM MDR Savings ($M)", [round(millions(RESULTS[n]["apm_mdr_savings"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_data(r, "APMs BENEFIT ($M)", [round(millions(RESULTS[n]["apms_benefit"]), 4) for n in countries], fmt="#,##0.00")
    for c_idx in range(2, 3+len(countries)+1):
        ws.cell(row=r, column=c_idx).font = TOTAL_FONT; ws.cell(row=r, column=c_idx).fill = TOTAL_FILL
    r += 2

    row_data(r, "TOTAL BENEFIT ($M)", [round(millions(RESULTS[n]["total_benefit"]), 2) for n in countries], fmt="#,##0.00")
    for c_idx in range(2, 3+len(countries)+1):
        ws.cell(row=r, column=c_idx).font = Font(bold=True, size=12); ws.cell(row=r, column=c_idx).fill = TOTAL_FILL

    ws.column_dimensions["B"].width = 34
    for i in range(len(countries)):
        ws.column_dimensions[get_column_letter(3+i)].width = 22
    ws.column_dimensions[get_column_letter(3+len(countries))].width = 18


def build_eng(region_name, countries):
    ws = wb.create_sheet(f"BC Devs.Eng {region_name}")
    ws["B1"] = f"BC DEVS/ENG — {region_name}"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF"); ws["B1"].fill = HDR_FILL
    last_col = get_column_letter(2 + len(countries) + 1)
    ws.merge_cells(f"B1:{last_col}1")

    ws["B3"] = "ENGINEERING INPUTS"
    for i, n in enumerate(countries):
        ws.cell(row=3, column=3+i, value=n)
    ws.cell(row=3, column=3+len(countries), value="TOTAL")
    for c in range(2, 3+len(countries)+1):
        ws.cell(row=3, column=c).font = HDR_FONT; ws.cell(row=3, column=c).fill = HDR_FILL

    def row(r, label, values, totalize=True, fmt=None):
        ws.cell(row=r, column=2, value=label)
        total = 0
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=3+i, value=v)
            if fmt and isinstance(v, (int, float)): cell.number_format = fmt
            if totalize and isinstance(v, (int, float)): total += v
        if totalize and isinstance(values[0], (int, float)):
            c = ws.cell(row=r, column=3+len(countries), value=round(total, 4))
            if fmt: c.number_format = fmt
            c.font = Font(bold=True)

    r = 4
    row(r, "APMs to Integrate (#)", [ENG_RESULTS[n]["n_apms"] for n in countries], fmt="0"); r += 1
    row(r, "APM Names", [", ".join(COUNTRIES[n]["to_integrate"]) for n in countries], totalize=False)
    for i in range(len(countries)):
        ws.cell(row=r, column=3+i).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 45
    r += 1
    row(r, "Time per APM Integration (months)", [TIME_APM]*len(countries), totalize=False); r += 1
    row(r, "APM Integration Months", [ENG_RESULTS[n]["apm_months"] for n in countries], fmt="0"); r += 1
    row(r, "APM Integration Cost ($M)", [round(millions(ENG_RESULTS[n]["apm_cost"]), 4) for n in countries], fmt="#,##0.0000"); r += 2

    row(r, "MoR/Processors to Integrate (#)", [ENG_RESULTS[n]["n_mors"] for n in countries], fmt="0"); r += 1
    row(r, "Need MoR?", [ENG_RESULTS[n]["need_mor"] for n in countries], totalize=False); r += 1
    row(r, "Time per PSP Integration (months)", [TIME_PSP]*len(countries), totalize=False); r += 1
    row(r, "PSP Integration Months", [ENG_RESULTS[n]["psp_months"] for n in countries], fmt="0"); r += 1
    row(r, "PSP Integration Cost ($M)", [round(millions(ENG_RESULTS[n]["psp_cost"]), 4) for n in countries], fmt="#,##0.0000"); r += 2

    row(r, "Total Integration Months", [ENG_RESULTS[n]["total_months"] for n in countries], fmt="0"); r += 1
    row(r, "Yuno Integration (months)", [YUNO_MONTHS]*len(countries), totalize=False); r += 1
    row(r, "TOTAL ENGINEERING SAVINGS ($M)", [round(millions(ENG_RESULTS[n]["total_savings"]), 4) for n in countries], fmt="#,##0.0000")
    for c_idx in range(2, 3+len(countries)+1):
        ws.cell(row=r, column=c_idx).font = TOTAL_FONT; ws.cell(row=r, column=c_idx).fill = TOTAL_FILL
    r += 2

    region_total_months = sum(ENG_RESULTS[n]["total_months"] for n in countries)
    ws.cell(row=r, column=2, value="Total months (all countries)"); ws.cell(row=r, column=3, value=region_total_months); r += 1
    ws.cell(row=r, column=2, value="Yuno integration (months)"); ws.cell(row=r, column=3, value=YUNO_MONTHS); r += 1
    ws.cell(row=r, column=2, value="Time Saved (months)"); ws.cell(row=r, column=3, value=region_total_months - YUNO_MONTHS)

    ws.column_dimensions["B"].width = 36
    for i in range(len(countries)):
        ws.column_dimensions[get_column_letter(3+i)].width = 22
    ws.column_dimensions[get_column_letter(3+len(countries))].width = 16


# Build pay-ins + eng for each region
REGION_ORDER = ["NA", "EMEA", "APAC"]
for reg in REGION_ORDER:
    build_payins(reg, REGIONS[reg])
for reg in REGION_ORDER:
    build_eng(reg, REGIONS[reg])


# ===== CURRENT PAYMENT METHODS =====
ws = wb.create_sheet("Current Payment Methods")
ws["B1"] = "CURRENT PAYMENT METHODS — CETTIRE"
ws["B1"].font = Font(bold=True, size=14, color="FFFFFF"); ws["B1"].fill = HDR_FILL
ws.merge_cells("B1:D1")

ws["B3"] = "Country"; ws["C3"] = "Current Payment Methods"; ws["D3"] = "Missing / Recommended APMs"
for c in "BCD":
    ws[f"{c}3"].font = HDR_FONT; ws[f"{c}3"].fill = HDR_FILL

r = 4
for region in REGION_ORDER:
    ws.cell(row=r, column=2, value=region).font = SECTION_FONT
    ws.cell(row=r, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    for name in REGIONS[region]:
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=CURRENT_PMS[name])
        ws.cell(row=r, column=4, value=", ".join(COUNTRIES[name]["to_integrate"]))
        for c in "BCD":
            ws[f"{c}{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 38
        r += 1

r += 2
notes = [
    "1. Platform confirmed as Shopify Plus (handle 4699f7-2.myshopify.com per Merchant Genius). Card acquiring runs through Shopify Payments (Stripe under the hood). No public evidence of Adyen, Braintree, Checkout.com, Worldpay, or Cybersource integrations.",
    "2. PayPal is NOT accepted in any market (verified via Knoji, whoacceptsit.com, customer reviews).",
    "3. Confirmed local legal entities: Australia (parent, ASX:CTT) and UK (Cettire Limited, Co. No. 10355981, Ipswich). All other 11 markets process through cross-border acquiring from the AU merchant of record.",
    "4. AOV anchors: Cettire-reported $798 (FY24), $825 (FY25 trading update), $961 (H1 FY26, Dec-2025). Per-market estimates adjusted for PPP and luxury-spend skew (Gulf and Hong Kong skew higher; Mexico lower).",
    "5. Customer sentiment: Trustpilot 1.5/5, ProductReview.com.au 1.5/5 (196 reviews), PissedConsumer 1.5/5 with 222 one-star reviews. Recurring complaints: refund delays, returns rejected on cosmetic grounds (forced into store credit), chargebacks cited as the only working resolution path.",
    "6. AFR exposé (Mar-2024) flagged duty-remittance opacity, dropped ASX:CTT 27% in one day. Cettire subsequently stopped showing duties as a separate line and rolled them into displayed price.",
    "7. US de minimis repeal (effective 29-Aug-2025) eliminated the $800 exemption — Cettire disclosed ~40% of May-Jun 2025 US gross revenue went below threshold. Direct hit to unit economics on the largest single market.",
    "8. China launched Jun-2024 via JD.com partnership (not in BC scope — not in SimilarWeb top traffic). Kuwait and Bahrain added FY25. Emerging markets now ~37% of gross revenue.",
    "9. APM recommendations sourced from the BC reference APMs tab (Canva BC) filtered to crucial local methods per market: iDEAL (NL 70% e-com), Mada (SA 93% domestic), KakaoPay/Naver Pay (KR ~25%), Konbini/PayPay (JP ~20%), Tabby/Tamara (UAE/SA ~8% BNPL), Interac (CA ~20%), OXXO/SPEI (MX ~25%), PayPal (US/UK/DE ~15-28%).",
    "10. All financials and customer counts pulled from FY24/FY25 ASX filings, listcorp and Capital Brief reporting. AUD→USD conversion at ~0.65. According to web data; assumptions made where entity registration could not be confirmed.",
]
for note in notes:
    ws.cell(row=r, column=2, value="NOTE")
    ws.cell(row=r, column=3, value=note)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 50
    r += 1

for col, w in [("B", 18), ("C", 70), ("D", 60)]:
    ws.column_dimensions[col].width = w

# Save
out = "Business Cases/BC sheets - Cettire.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"Total benefit: ${total:.2f}M | Acc uplift: ${acc_uplift:.2f}M | Fee reno: ${fee_reno:.2f}M | APM growth: ${apm_growth:.2f}M | Eng savings: ${eng_savings_m:.2f}M")
print(f"Integration months: {TOTAL_INTEGRATION_MONTHS} | Time saved: {TIME_SAVED_MONTHS} mo ({TIME_SAVED_MONTHS/12:.1f} yr)")
