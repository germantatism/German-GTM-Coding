"""Build BC sheets - GOAT.xlsx mirroring the Nordstrom/Abercrombie/Cettire template."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- INPUTS ----------
ANNUAL_VISITS = 157_100_000
CONVERSION = 0.06
VTC = 0.20   # Marketplace benchmark
CTS = 0.80   # Marketplace benchmark
TAKE_RATE = 0.12  # GOAT blended take rate (9.5%-25% seller commission + 2.9% cash-out; effective ~12-15%)
APM_MDR_COST = 0.012
MONTHS_COST = 21_600
TIME_APM = 3
TIME_PSP = 4
YUNO_MONTHS = 2  # per region

# Industry benchmark — Marketplace
LOCAL_MDR = 0.025
CB_MDR = 0.055
DELTA_LOCAL = 0.05
DELTA_CB = 0.10

COUNTRIES = {
    # NA
    "United States":  {"region": "NA",   "sow": 0.7578, "atv": 310, "entity": "Yes", "current_apms": 8, "to_integrate": ["Venmo", "Cash App Pay"],                                                  "sow_cards": 0.70, "pct_inc": 0.03, "pct_shift": 0.02},
    "Canada":         {"region": "NA",   "sow": 0.0322, "atv": 295, "entity": "No",  "current_apms": 5, "to_integrate": ["Interac e-Transfer", "Interac Online"],                                  "sow_cards": 0.82, "pct_inc": 0.08, "pct_shift": 0.04},
    "Mexico":         {"region": "NA",   "sow": 0.0110, "atv": 185, "entity": "No",  "current_apms": 3, "to_integrate": ["OXXO Pay", "SPEI", "Mercado Pago", "MSI Installments"],                "sow_cards": 0.90, "pct_inc": 0.12, "pct_shift": 0.05},
    # EMEA
    "United Kingdom": {"region": "EMEA", "sow": 0.0553, "atv": 320, "entity": "Yes", "current_apms": 6, "to_integrate": ["Open Banking", "BACS Direct Debit"],                                     "sow_cards": 0.72, "pct_inc": 0.04, "pct_shift": 0.02},
    "France":         {"region": "EMEA", "sow": 0.0232, "atv": 300, "entity": "Yes", "current_apms": 4, "to_integrate": ["Carte Bancaire", "Wero (ex-Paylib)", "Alma BNPL"],                       "sow_cards": 0.85, "pct_inc": 0.10, "pct_shift": 0.05},
}

CURRENT_PMS = {
    "United States":  "Visa, Mastercard, Amex, Discover, JCB; Apple Pay, Google Pay, PayPal, Alipay; Affirm, Klarna, Afterpay, Zip; GOAT Gift Card, GOAT Credit",
    "Canada":         "Visa, Mastercard, Amex; Apple Pay, Google Pay, PayPal; Klarna, Afterpay, Zip",
    "Mexico":         "Visa, Mastercard, Amex; Apple Pay, Google Pay, PayPal",
    "United Kingdom": "Visa, Mastercard, Amex; Apple Pay, Google Pay, PayPal; Klarna, Clearpay (Afterpay UK), Zip",
    "France":         "Visa, Mastercard, Amex; Apple Pay, Google Pay, PayPal; Klarna, giropay (DE-leak)",
}


def compute(c):
    visits = ANNUAL_VISITS * c["sow"]
    txns = visits * CONVERSION
    tpv = txns * c["atv"]
    tpv_cards = tpv * c["sow_cards"]
    sow_apms = 1 - c["sow_cards"]
    tpv_apms = tpv * sow_apms
    delta = DELTA_LOCAL if c["entity"] == "Yes" else DELTA_CB
    inc_tpv = tpv_cards * delta
    inc_rev = inc_tpv * TAKE_RATE
    annual_cards_tpv = tpv_cards + inc_tpv
    current_mdr = LOCAL_MDR if c["entity"] == "Yes" else CB_MDR
    mdr_reduction = -0.002 if c["entity"] == "Yes" else -0.01
    new_mdr = current_mdr + mdr_reduction
    mdr_savings = annual_cards_tpv * abs(mdr_reduction)
    cards_benefit = inc_rev + mdr_savings

    apm_inc_tpv = tpv * c["pct_inc"]
    apm_inc_rev = apm_inc_tpv * TAKE_RATE
    apm_shifted_tpv = tpv * c["pct_shift"]
    apm_mdr_savings = apm_shifted_tpv * (current_mdr - APM_MDR_COST)
    apms_benefit = apm_inc_rev + apm_mdr_savings

    total_benefit = cards_benefit + apms_benefit

    return {
        "visits": visits, "txns": txns, "tpv": tpv,
        "tpv_cards": tpv_cards, "sow_apms": sow_apms, "tpv_apms": tpv_apms,
        "delta": delta, "inc_tpv": inc_tpv, "inc_rev": inc_rev,
        "annual_cards_tpv": annual_cards_tpv,
        "current_mdr": current_mdr, "mdr_reduction": mdr_reduction, "new_mdr": new_mdr,
        "mdr_savings": mdr_savings, "cards_benefit": cards_benefit,
        "apm_inc_tpv": apm_inc_tpv, "apm_inc_rev": apm_inc_rev,
        "apm_shifted_tpv": apm_shifted_tpv, "apm_mdr_savings": apm_mdr_savings,
        "apms_benefit": apms_benefit, "total_benefit": total_benefit,
    }


RESULTS = {name: compute(c) for name, c in COUNTRIES.items()}

REGIONS = {"NA": [], "EMEA": []}
for name, c in COUNTRIES.items():
    REGIONS[c["region"]].append(name)

def millions(x): return x / 1_000_000

TOTAL_INC_REV = sum(r["inc_rev"] for r in RESULTS.values())
TOTAL_MDR_SAVINGS = sum(r["mdr_savings"] for r in RESULTS.values())
TOTAL_APM_INC_REV = sum(r["apm_inc_rev"] for r in RESULTS.values())
TOTAL_APM_MDR_SAVINGS = sum(r["apm_mdr_savings"] for r in RESULTS.values())
TOTAL_APM_BENEFIT = TOTAL_APM_INC_REV + TOTAL_APM_MDR_SAVINGS

def eng_for_country(name):
    c = COUNTRIES[name]
    n_apms = len(c["to_integrate"])
    apm_months = n_apms * TIME_APM
    apm_cost = apm_months * MONTHS_COST
    n_mors = 2 if c["entity"] == "No" else 1
    need_mor = "Yes" if c["entity"] == "No" else "No"
    psp_months = n_mors * TIME_PSP
    psp_cost = psp_months * MONTHS_COST
    total_months = apm_months + psp_months
    total_savings = apm_cost + psp_cost
    return {
        "n_apms": n_apms, "apm_months": apm_months, "apm_cost": apm_cost,
        "n_mors": n_mors, "need_mor": need_mor, "psp_months": psp_months, "psp_cost": psp_cost,
        "total_months": total_months, "total_savings": total_savings,
    }

ENG_RESULTS = {name: eng_for_country(name) for name in COUNTRIES}
TOTAL_ENG_SAVINGS = sum(e["total_savings"] for e in ENG_RESULTS.values())
TOTAL_INTEGRATION_MONTHS = sum(e["total_months"] for e in ENG_RESULTS.values())
YUNO_TOTAL_MONTHS = YUNO_MONTHS * 2  # 2 regions
TIME_SAVED_MONTHS = TOTAL_INTEGRATION_MONTHS - YUNO_TOTAL_MONTHS

# ---------- BUILD WORKBOOK ----------
wb = openpyxl.Workbook()
wb.remove(wb.active)

HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
SECTION_FONT = Font(bold=True, color="111827", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="FEF3C7")
TOTAL_FONT = Font(bold=True, size=11)

# ===== RESUMEN =====
ws = wb.create_sheet("RESUMEN")
ws["B1"] = "BUSINESS CASE — GOAT"
ws["B1"].font = Font(bold=True, size=16, color="FFFFFF"); ws["B1"].fill = HDR_FILL
ws.merge_cells("B1:H1")

ws["B3"] = "Company";              ws["C3"] = "GOAT Group (1661, Inc.)"
ws["B4"] = "Industry";             ws["C4"] = "Marketplace (sneakers + streetwear + luxury resale)"
ws["B5"] = "Annual Total Visits";  ws["C5"] = ANNUAL_VISITS
ws["B6"] = "Annual Revenue ($M)";  ws["C6"] = 248
ws["B7"] = "Take Rate";            ws["C7"] = TAKE_RATE

ws["B9"] = "TOTAL BENEFIT SUMMARY"
ws["B9"].font = SECTION_FONT; ws["B9"].fill = SECTION_FILL
ws.merge_cells("B9:E9")

ws["B10"] = "Category"; ws["C10"] = "Value ($M)"; ws["D10"] = "Lower Range ($M)"; ws["E10"] = "Upper Range ($M)"
for c in "BCDE":
    ws[f"{c}10"].font = HDR_FONT; ws[f"{c}10"].fill = HDR_FILL

acc_uplift = millions(TOTAL_INC_REV)
fee_reno = millions(TOTAL_MDR_SAVINGS)
apm_growth = millions(TOTAL_APM_BENEFIT)
eng_savings_m = millions(TOTAL_ENG_SAVINGS)
total = acc_uplift + fee_reno + apm_growth + eng_savings_m

rows = [
    ("TPV Acceptance Rate Uplift",     acc_uplift),
    ("Fee Renegotiation (MDR Savings)", fee_reno),
    ("New APM TPV Growth",              apm_growth),
    ("Engineering Cost Savings",        eng_savings_m),
    ("TOTAL BENEFIT",                   total),
]
for i, (label, val) in enumerate(rows, start=11):
    ws[f"B{i}"] = label
    ws[f"C{i}"] = round(val, 2)
    ws[f"D{i}"] = round(val * 0.85, 2)
    ws[f"E{i}"] = round(val * 1.15, 2)
    if label == "TOTAL BENEFIT":
        for c in "BCDE":
            ws[f"{c}{i}"].font = TOTAL_FONT; ws[f"{c}{i}"].fill = TOTAL_FILL

ws["B17"] = "Total integration months (without Yuno)"; ws["C17"] = TOTAL_INTEGRATION_MONTHS
ws["B18"] = "Yuno integration time (months, per region x2)"; ws["C18"] = YUNO_TOTAL_MONTHS
ws["B19"] = "Time Saved (months)";                     ws["C19"] = TIME_SAVED_MONTHS
ws["B20"] = "Time Saved (years)";                      ws["C20"] = round(TIME_SAVED_MONTHS / 12, 1)

problems = [
    ("PROBLEM 1", "Multi-PSP patchwork likely anchored on Braintree (PayPal-owned) with 6+ direct APM contracts (Affirm, Klarna, Afterpay, Zip, Alipay, giropay); no public orchestration, smart routing, or multi-acquirer failover. BNPL stripped from Drops and Auctions as a blunt fraud control."),
    ("PROBLEM 2", "FTC court order Dec 2024: $2,013,527 consumer refund for Mail Order Rule violations and Buyer Protection failures (37% of Instant orders shipped late, 16% of Next Day late, partial refunds and forced store-credit on inauthentic items)."),
    ("PROBLEM 3", "Refund-as-store-credit pattern triggers chargebacks; BBB documents 1,647 complaints in 3 years (510 in last 12 months); GOAT applies a 90-day refund freeze once a chargeback is filed, extending dispute resolution and inflating dispute costs."),
    ("PROBLEM 4", "France market activation 2025 (WPP Media-led) but checkout has no Carte Bancaire native routing; ~35% of French online cards co-badged with CB hit foreign-acquirer rails with higher decline rates and avoidable interchange."),
    ("PROBLEM 5", "Missing local rails across NA/EMEA: no Interac e-Transfer/Online in Canada (dominant consumer debit), no OXXO/SPEI/Mercado Pago/MSI installments in Mexico (entire sub-$500 sneaker buyer segment), no Venmo/Cash App Pay in US gen-Z core."),
]
for i, (k, v) in enumerate(problems, start=22):
    ws[f"G{i}"] = k; ws[f"H{i}"] = v
    ws[f"G{i}"].font = Font(bold=True)
    ws[f"H{i}"].alignment = Alignment(wrap_text=True, vertical="top")

for col, w in [("B", 38), ("C", 22), ("D", 20), ("E", 20), ("G", 12), ("H", 110)]:
    ws.column_dimensions[col].width = w
for r in range(22, 27):
    ws.row_dimensions[r].height = 60


# ===== PAY-INS SHEETS =====
def build_payins(region_name, countries):
    ws = wb.create_sheet(f"BC Pay-ins {region_name}")
    ws[f"B1"] = f"BC PAY-INS — {region_name}"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF"); ws["B1"].fill = HDR_FILL
    last_col = get_column_letter(2 + len(countries) + 1)
    ws.merge_cells(f"B1:{last_col}1")

    ws["B3"] = "SDR INPUTS"
    for i, name in enumerate(countries):
        ws.cell(row=3, column=3+i, value=name)
    ws.cell(row=3, column=3+len(countries), value="TOTAL")
    for c in range(2, 3+len(countries)+1):
        ws.cell(row=3, column=c).font = HDR_FONT; ws.cell(row=3, column=c).fill = HDR_FILL

    def row_section(r, label):
        ws.cell(row=r, column=2, value=label).font = SECTION_FONT
        ws.cell(row=r, column=2).fill = SECTION_FILL

    def row_data(r, label, values, totalize=True, fmt=None):
        ws.cell(row=r, column=2, value=label)
        total = 0
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=3+i, value=v)
            if isinstance(v, (int, float)) and fmt:
                cell.number_format = fmt
            if isinstance(v, (int, float)) and totalize:
                total += v
        if totalize and isinstance(values[0], (int, float)):
            c = ws.cell(row=r, column=3+len(countries), value=round(total, 4))
            if fmt: c.number_format = fmt
            c.font = Font(bold=True)

    r = 4
    row_section(r, "TRAFFIC & TPV"); r += 1
    row_data(r, "Web Traffic (SoW)", [round(COUNTRIES[n]["sow"], 4) for n in countries], totalize=False)
    for i,n in enumerate(countries):
        ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Web Traffic Visits", [int(RESULTS[n]["visits"]) for n in countries], fmt="#,##0"); r += 1
    row_data(r, "Conversion", [CONVERSION]*len(countries), totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "View to Checkout (VtC)", [VTC]*len(countries), totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Checkout to Success (CtS)", [CTS]*len(countries), totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Successful Transactions/Year", [int(RESULTS[n]["txns"]) for n in countries], fmt="#,##0"); r += 1
    row_data(r, "ATV ($)", [COUNTRIES[n]["atv"] for n in countries], totalize=False); r += 1
    row_data(r, "TPV ($M)", [round(millions(RESULTS[n]["tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "SoW Cards (%)", [COUNTRIES[n]["sow_cards"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "TPV Cards ($M)", [round(millions(RESULTS[n]["tpv_cards"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "SoW APMs (%)", [round(RESULTS[n]["sow_apms"], 4) for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "TPV APMs ($M)", [round(millions(RESULTS[n]["tpv_apms"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_section(r, "CARDS BENEFIT"); r += 1
    row_data(r, "Delta Acceptance Rate (%)", [RESULTS[n]["delta"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Incremental TPV ($M)", [round(millions(RESULTS[n]["inc_tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "Incremental Revenue ($M)", [round(millions(RESULTS[n]["inc_rev"]), 4) for n in countries], fmt="#,##0.00"); r += 2
    row_data(r, "Annual Cards TPV ($M)", [round(millions(RESULTS[n]["annual_cards_tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "Local Entity?", [COUNTRIES[n]["entity"] for n in countries], totalize=False); r += 1
    row_data(r, "Current MDR", [RESULTS[n]["current_mdr"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "Reduction in MDR", [RESULTS[n]["mdr_reduction"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "New MDR", [RESULTS[n]["new_mdr"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "MDR Reduction Savings ($M)", [round(millions(RESULTS[n]["mdr_savings"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_data(r, "CARDS BENEFIT ($M)", [round(millions(RESULTS[n]["cards_benefit"]), 4) for n in countries], fmt="#,##0.00")
    for c_idx in range(2, 3+len(countries)+1):
        ws.cell(row=r, column=c_idx).font = TOTAL_FONT; ws.cell(row=r, column=c_idx).fill = TOTAL_FILL
    r += 2

    row_section(r, "APMs BENEFIT"); r += 1
    row_data(r, "APMs to Integrate", [", ".join(COUNTRIES[n]["to_integrate"]) for n in countries], totalize=False)
    for i in range(len(countries)):
        ws.cell(row=r, column=3+i).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 45
    r += 1
    row_data(r, "Current APMs (#)", [COUNTRIES[n]["current_apms"] for n in countries], fmt="0"); r += 1
    row_data(r, "APMs to Integrate (#)", [len(COUNTRIES[n]["to_integrate"]) for n in countries], fmt="0"); r += 2

    row_data(r, "(%) Incremental Transactions", [COUNTRIES[n]["pct_inc"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "APMs Incremental TPV ($M)", [round(millions(RESULTS[n]["apm_inc_tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 1
    row_data(r, "APMs Incremental Revenue ($M)", [round(millions(RESULTS[n]["apm_inc_rev"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_data(r, "(%) TPV from Cards to APMs", [COUNTRIES[n]["pct_shift"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "APMs Shifted TPV ($M)", [round(millions(RESULTS[n]["apm_shifted_tpv"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_data(r, "Cards MDR", [RESULTS[n]["current_mdr"] for n in countries], totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "APMs MDR Cost", [APM_MDR_COST]*len(countries), totalize=False)
    for i in range(len(countries)): ws.cell(row=r, column=3+i).number_format = "0.00%"
    r += 1
    row_data(r, "APM MDR Savings ($M)", [round(millions(RESULTS[n]["apm_mdr_savings"]), 4) for n in countries], fmt="#,##0.00"); r += 2

    row_data(r, "APMs BENEFIT ($M)", [round(millions(RESULTS[n]["apms_benefit"]), 4) for n in countries], fmt="#,##0.00")
    for c_idx in range(2, 3+len(countries)+1):
        ws.cell(row=r, column=c_idx).font = TOTAL_FONT; ws.cell(row=r, column=c_idx).fill = TOTAL_FILL
    r += 2

    row_data(r, "TOTAL BENEFIT ($M)", [round(millions(RESULTS[n]["total_benefit"]), 2) for n in countries], fmt="#,##0.00")
    for c_idx in range(2, 3+len(countries)+1):
        ws.cell(row=r, column=c_idx).font = Font(bold=True, size=12); ws.cell(row=r, column=c_idx).fill = TOTAL_FILL

    ws.column_dimensions["B"].width = 34
    for i in range(len(countries)):
        ws.column_dimensions[get_column_letter(3+i)].width = 22
    ws.column_dimensions[get_column_letter(3+len(countries))].width = 18


def build_eng(region_name, countries):
    ws = wb.create_sheet(f"BC Devs.Eng {region_name}")
    ws["B1"] = f"BC DEVS/ENG — {region_name}"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF"); ws["B1"].fill = HDR_FILL
    last_col = get_column_letter(2 + len(countries) + 1)
    ws.merge_cells(f"B1:{last_col}1")

    ws["B3"] = "ENGINEERING INPUTS"
    for i, n in enumerate(countries):
        ws.cell(row=3, column=3+i, value=n)
    ws.cell(row=3, column=3+len(countries), value="TOTAL")
    for c in range(2, 3+len(countries)+1):
        ws.cell(row=3, column=c).font = HDR_FONT; ws.cell(row=3, column=c).fill = HDR_FILL

    def row(r, label, values, totalize=True, fmt=None):
        ws.cell(row=r, column=2, value=label)
        total = 0
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=3+i, value=v)
            if fmt and isinstance(v, (int, float)): cell.number_format = fmt
            if totalize and isinstance(v, (int, float)): total += v
        if totalize and isinstance(values[0], (int, float)):
            c = ws.cell(row=r, column=3+len(countries), value=round(total, 4))
            if fmt: c.number_format = fmt
            c.font = Font(bold=True)

    r = 4
    row(r, "APMs to Integrate (#)", [ENG_RESULTS[n]["n_apms"] for n in countries], fmt="0"); r += 1
    row(r, "APM Names", [", ".join(COUNTRIES[n]["to_integrate"]) for n in countries], totalize=False)
    for i in range(len(countries)):
        ws.cell(row=r, column=3+i).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 45
    r += 1
    row(r, "Time per APM Integration (months)", [TIME_APM]*len(countries), totalize=False); r += 1
    row(r, "APM Integration Months", [ENG_RESULTS[n]["apm_months"] for n in countries], fmt="0"); r += 1
    row(r, "APM Integration Cost ($M)", [round(millions(ENG_RESULTS[n]["apm_cost"]), 4) for n in countries], fmt="#,##0.0000"); r += 2

    row(r, "MoR/Processors to Integrate (#)", [ENG_RESULTS[n]["n_mors"] for n in countries], fmt="0"); r += 1
    row(r, "Need MoR?", [ENG_RESULTS[n]["need_mor"] for n in countries], totalize=False); r += 1
    row(r, "Time per PSP Integration (months)", [TIME_PSP]*len(countries), totalize=False); r += 1
    row(r, "PSP Integration Months", [ENG_RESULTS[n]["psp_months"] for n in countries], fmt="0"); r += 1
    row(r, "PSP Integration Cost ($M)", [round(millions(ENG_RESULTS[n]["psp_cost"]), 4) for n in countries], fmt="#,##0.0000"); r += 2

    row(r, "Total Integration Months", [ENG_RESULTS[n]["total_months"] for n in countries], fmt="0"); r += 1
    row(r, "Yuno Integration (months)", [YUNO_MONTHS]*len(countries), totalize=False); r += 1
    row(r, "TOTAL ENGINEERING SAVINGS ($M)", [round(millions(ENG_RESULTS[n]["total_savings"]), 4) for n in countries], fmt="#,##0.0000")
    for c_idx in range(2, 3+len(countries)+1):
        ws.cell(row=r, column=c_idx).font = TOTAL_FONT; ws.cell(row=r, column=c_idx).fill = TOTAL_FILL
    r += 2

    region_total_months = sum(ENG_RESULTS[n]["total_months"] for n in countries)
    ws.cell(row=r, column=2, value="Total months (all countries)"); ws.cell(row=r, column=3, value=region_total_months); r += 1
    ws.cell(row=r, column=2, value="Yuno integration (months)"); ws.cell(row=r, column=3, value=YUNO_MONTHS); r += 1
    ws.cell(row=r, column=2, value="Time Saved (months)"); ws.cell(row=r, column=3, value=region_total_months - YUNO_MONTHS)

    ws.column_dimensions["B"].width = 36
    for i in range(len(countries)):
        ws.column_dimensions[get_column_letter(3+i)].width = 22
    ws.column_dimensions[get_column_letter(3+len(countries))].width = 16


REGION_ORDER = ["NA", "EMEA"]
for reg in REGION_ORDER:
    build_payins(reg, REGIONS[reg])
for reg in REGION_ORDER:
    build_eng(reg, REGIONS[reg])


# ===== CURRENT PAYMENT METHODS =====
ws = wb.create_sheet("Current Payment Methods")
ws["B1"] = "CURRENT PAYMENT METHODS — GOAT"
ws["B1"].font = Font(bold=True, size=14, color="FFFFFF"); ws["B1"].fill = HDR_FILL
ws.merge_cells("B1:D1")

ws["B3"] = "Country"; ws["C3"] = "Current Payment Methods"; ws["D3"] = "Missing / Recommended APMs"
for c in "BCD":
    ws[f"{c}3"].font = HDR_FONT; ws[f"{c}3"].fill = HDR_FILL

r = 4
for region in REGION_ORDER:
    ws.cell(row=r, column=2, value=region).font = SECTION_FONT
    ws.cell(row=r, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    for name in REGIONS[region]:
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=CURRENT_PMS[name])
        ws.cell(row=r, column=4, value=", ".join(COUNTRIES[name]["to_integrate"]))
        for c in "BCD":
            ws[f"{c}{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 38
        r += 1

r += 2
notes = [
    "1. PSP identification inferred from public behavior, GOAT job postings (Associate Director, Payments role calling for multi-acquirer experience), and ecosystem signals. Braintree (PayPal-owned) is the most likely card-acquiring anchor given native PayPal, Apple Pay, Google Pay, tokenized card storage. Affirm, Klarna, Afterpay, Zip, Alipay, giropay are confirmed direct integrations per GOAT Support. Stripe or Adyen possible as secondary rails. According to web data; not publicly disclosed by GOAT.",
    "2. Confirmed legal entities: US (1661, Inc., Delaware/CA, d/b/a GOAT), UK (GOAT Group Holdings Limited, Co. #13883004, Altrincham), France (GOAT SAS, registered with AMF). Canada and Mexico have NO confirmed local entity — Canadian sellers are priced separately (12.4% commission vs. 9.5% US baseline), but checkout still acquires cross-border.",
    "3. FTC court order December 2024: $2,013,527 consumer refund required for Mail Order Rule violations and Buyer Protection failures. Findings: 37% of Instant orders shipped late, 16% of Next Day orders late, partial refunds and forced store-credit on items GOAT deemed inauthentic.",
    "4. Customer sentiment: BBB 1,647 complaints over 3 years (510 in last 12 months); Sitejabber 1.1/5 (200 reviews); Trustpilot 4.0/5 (40K reviews, bimodal). Recurring complaints: refunds routed to GOAT Credit instead of original tender, 90-day refund freeze once a chargeback is filed, counterfeit/SNAD disputes refused despite photo evidence.",
    "5. ATV anchors: GOAT-reported AOV $300+ post-Affirm integration (+30% lift vs pre-Affirm mid-$200s). Per-market estimates adjusted for PPP (Mexico ~$185, UK/FR ~$300-320) and category mix (luxury via Grailed pulls UK/FR higher; Sneakers.com sub-brand at ~$70 AOV excluded as separate platform).",
    "6. International consolidation Sept 2025: GOAT closed authentication facilities in Hong Kong, Japan, and the UK; wound down alias platform internationally (seller inventory returned Oct 2025). France market activation 2025 (WPP Media performance + influencer push) is the active growth bet.",
    "7. APM recommendations sourced from the BC reference APMs tab filtered to crucial local methods per market: Carte Bancaire (FR ~35% of online cards, co-badged with Visa/MC), Interac e-Transfer/Online (CA ~20% of online), OXXO Pay/SPEI/Mercado Pago/MSI installments (MX combined ~25-30% of e-com), Venmo + Cash App Pay (US gen-Z core).",
    "8. Take rate: GOAT charges sellers 9.5% baseline (sellers rated 90+) scaling to 25% for sub-50 rated, plus 2.9% cash-out fee and $5-$30 fixed per-item seller fee. Canadian sellers priced at 12.4%/17.9%/22.9%/27.9%. Blended effective take rate used in BC: 12% — midpoint of 10-15% range implied by mix.",
    "9. Cross-border: GOAT is buyer-facing MoR; sellers paid in USD (0.6% cross-border fee if cashing out in USD from outside US). DDP for AU, CA, MX, EU (duties collected at checkout). DDU for other destinations — buyer pays carrier at delivery with full price declared for customs.",
    "10. All financials cross-checked against GOAT Group press releases, FTC case page, Companies House (UK), AMF France, Bloomberg 1576186D:US profile, Latka self-reported revenue, and Tracxn/Crunchbase funding history. According to web data; assumptions made where formal disclosure unavailable.",
]
for note in notes:
    ws.cell(row=r, column=2, value="NOTE")
    ws.cell(row=r, column=3, value=note)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 55
    r += 1

for col, w in [("B", 18), ("C", 70), ("D", 60)]:
    ws.column_dimensions[col].width = w

out = "Business Cases/BC sheets - GOAT.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"Total benefit: ${total:.2f}M | Acc uplift: ${acc_uplift:.2f}M | Fee reno: ${fee_reno:.2f}M | APM growth: ${apm_growth:.2f}M | Eng savings: ${eng_savings_m:.2f}M")
print(f"Integration months: {TOTAL_INTEGRATION_MONTHS} | Time saved: {TIME_SAVED_MONTHS} mo ({TIME_SAVED_MONTHS/12:.1f} yr)")
